"""
Smoke test — verifies every layer of the V1 pipeline without running the LP solve
or making any DB writes. Run from project root:

    python3 smoke_test.py

Pass = exit 0. Fail = exit 1, with the first failed check named.
"""
from __future__ import annotations

import importlib
import sys
import traceback
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))


def _discover_test_plan_id() -> str | None:
    """Pick any plan_id that has rows in BOTH jkt_demand AND jkt_plan_params.
    Returns None if the DB has no usable test plan (smoke checks that need
    demand will print 'skipped' instead of failing)."""
    from V1.utilities import config_loader, db
    try:
        cfg = config_loader.load()
        conn = db.connect(cfg["db"])
        cur = conn.cursor()
        cur.execute("""
            SELECT d.plan_id
            FROM jkt_demand d
            JOIN jkt_plan_params p ON p.plan_id = d.plan_id
            GROUP BY d.plan_id
            HAVING COUNT(*) > 0
            LIMIT 1
        """)
        row = cur.fetchone()
        cur.close(); conn.close()
        return row[0] if row else None
    except Exception:
        return None


TEST_PLAN_ID = _discover_test_plan_id()

CHECKS: list[tuple[str, callable]] = []


def _skip_if_no_test_plan(name: str) -> bool:
    if TEST_PLAN_ID is None:
        print(f"  SKIP  {name}  (no plan_id has rows in both jkt_demand and jkt_plan_params)")
        return True
    return False


def check(name: str):
    def deco(fn):
        CHECKS.append((name, fn))
        return fn
    return deco


# --------------------------------------------------------------------------- #
# 1. Filesystem layout                                                        #
# --------------------------------------------------------------------------- #
@check("filesystem: required dirs exist")
def _():
    for d in ["V1", "V1/routes", "V1/utilities", "V1/setups", "V1/reports",
              "config", "input", "output"]:
        assert (ROOT / d).is_dir(), f"missing dir: {d}"


@check("filesystem: required files exist")
def _():
    for f in ["main.py", "app.py", "config/config.yaml",
              "V1/routes/demand_route.py", "V1/routes/schedule_route.py",
              "V1/routes/upload_route.py", "V1/routes/api_route.py",
              "V1/setups/plan_params.py", "V1/setups/demand_db.py",
              "V1/utilities/db.py", "V1/utilities/config_loader.py",
              "V1/reports/kpi_writer.py", "V1/reports/plan_writer.py",
              "V1/reports/capacity_writer.py"]:
        assert (ROOT / f).is_file(), f"missing file: {f}"


# --------------------------------------------------------------------------- #
# 2. Import surface                                                           #
# --------------------------------------------------------------------------- #
@check("imports: all V1 modules load cleanly")
def _():
    for mod in ["V1.utilities.config_loader", "V1.utilities.db",
                "V1.setups.plan_params", "V1.setups.demand_db",
                "V1.routes.demand_route", "V1.routes.schedule_route",
                "V1.routes.upload_route", "V1.routes.api_route",
                "V1.reports.kpi_writer", "V1.reports.plan_writer",
                "V1.reports.capacity_writer"]:
        importlib.import_module(mod)


# --------------------------------------------------------------------------- #
# 3. Config                                                                   #
# --------------------------------------------------------------------------- #
@check("config: loads and has required sections")
def _():
    from V1.utilities import config_loader
    cfg = config_loader.load()
    for k in ["plan", "db", "paths", "demand", "schedule", "upload"]:
        assert k in cfg, f"missing top-level config key: {k}"
    assert cfg["plan"]["plan_id"], "plan.plan_id is empty"


@check("config: paths resolve to absolute dirs")
def _():
    from V1.utilities import config_loader
    cfg = config_loader.load()
    assert config_loader.input_dir(cfg).is_dir()
    assert config_loader.output_dir(cfg).is_dir()


# --------------------------------------------------------------------------- #
# 4. DB connectivity & schema                                                 #
# --------------------------------------------------------------------------- #
@check("db: mysql.connector connect() works")
def _():
    from V1.utilities import config_loader, db
    cfg = config_loader.load()
    conn = db.connect(cfg["db"])
    cur = conn.cursor()
    cur.execute("SELECT 1")
    assert cur.fetchone()[0] == 1
    cur.close(); conn.close()


@check("db: sqlalchemy engine() works")
def _():
    from V1.utilities import config_loader, db
    import pandas as pd
    cfg = config_loader.load()
    eng = db.engine(cfg["db"])
    df = pd.read_sql("SELECT 1 AS x", eng)
    assert df.iloc[0]["x"] == 1


@check("db: all required tables exist")
def _():
    from V1.utilities import config_loader, db
    cfg = config_loader.load()
    conn = db.connect(cfg["db"]); cur = conn.cursor()
    needed = ["jkt_plan_params", "jkt_demand", "jkt_plan_kpis", "jkt_plan",
              "jkt_plan_capacityUtilisation", "Master_Curing_Design_CycleTime",
              "Master_Curing_Allowable_Machines_source", "gt_inventory_manual",
              "Master_WC_Master", "Daily_Running_Moulds",
              "Master_Mapping_Mould_SKU"]
    for t in needed:
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        cur.fetchone()
    cur.close(); conn.close()


# --------------------------------------------------------------------------- #
# 5. Setups                                                                   #
# --------------------------------------------------------------------------- #
@check("setups: plan_params.fetch returns expected fields")
def _():
    if _skip_if_no_test_plan("setups: plan_params.fetch returns expected fields"): return
    from V1.utilities import config_loader
    from V1.setups import plan_params
    cfg = config_loader.load()
    row = plan_params.fetch(cfg["db"], TEST_PLAN_ID)
    for k in ["planStartDate", "planEndDate", "oe", "re",
              "marketWeightage", "quantityWeightage", "targetdateWeightage"]:
        assert k in row, f"plan_params missing: {k}"


@check("setups: demand_db.load returns shaped rows")
def _():
    if _skip_if_no_test_plan("setups: demand_db.load returns shaped rows"): return
    from V1.utilities import config_loader
    from V1.setups import demand_db
    cfg = config_loader.load()
    rows = demand_db.load(cfg["db"], TEST_PLAN_ID)
    assert len(rows) > 0, "no demand rows for test plan"
    r = rows[0]
    for k in ["SKUCode", "SKU Description", "Requirement", "Market", "Delivery date"]:
        assert k in r, f"demand row missing key: {k}"


# --------------------------------------------------------------------------- #
# 6. Demand route logic (pure)                                                #
# --------------------------------------------------------------------------- #
@check("demand_route: market_score inverts rank correctly")
def _():
    from V1.routes.demand_route import _market_score
    plan = {"oe": 1, "re": 2, "government": 7}
    aliases = {"OE": "oe", "RE": "re", "Government": "government"}
    # rank 1 -> highest -> score 7. rank 7 -> lowest -> score 1.
    assert _market_score("OE", plan, aliases, 1, 7) == 7
    assert _market_score("RE", plan, aliases, 1, 7) == 6
    assert _market_score("Government", plan, aliases, 1, 7) == 1


@check("demand_route: end-to-end CPS for test plan (no Excel input needed)")
def _():
    if _skip_if_no_test_plan("demand_route: end-to-end CPS for test plan (no Excel input needed)"): return
    from V1.utilities import config_loader
    from V1.routes import demand_route
    cfg = config_loader.load()
    cfg["plan"]["plan_id"] = TEST_PLAN_ID
    cfg = config_loader.resolve_paths(cfg)
    out = demand_route.run(cfg)
    assert out.exists(), f"output file not written: {out}"
    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=6).value == "ConsolidatedPriorityScore"
    assert ws.max_row > 1


# --------------------------------------------------------------------------- #
# 7. Flask layer                                                              #
# --------------------------------------------------------------------------- #
@check("flask: app instantiates with expected routes")
def _():
    from app import create_app
    app = create_app()
    rules = {r.rule for r in app.url_map.iter_rules()}
    assert "/app/v1/jkt/planning-scheduling/plan/generate-plan" in rules
    assert "/app/v1/jkt/planning-scheduling/health" in rules


@check("flask: /health responds 200 via test client")
def _():
    from app import create_app
    client = create_app().test_client()
    r = client.get("/app/v1/jkt/planning-scheduling/health")
    assert r.status_code == 200, r.status_code
    assert r.get_json() == {"status": "ok"}


@check("flask: /generate- with missing plan_id returns 400")
def _():
    from app import create_app
    client = create_app().test_client()
    r = client.post("/app/v1/jkt/planning-scheduling/plan/generate-plan",
                    json={})
    assert r.status_code == 400, r.status_code
    assert r.get_json()["status"] == "error"


@check("flask: /generate- extracts plan_id correctly from JSON body")
def _():
    """Stop short of running the pipeline by using a plan_id known to be missing
    from jkt_plan_params — pipeline returns 404, but only after extracting."""
    from app import create_app
    client = create_app().test_client()
    r = client.post("/app/v1/jkt/planning-scheduling/plan/generate-plan",
                    json={"plan_id": "  __DOES_NOT_EXIST__  "})  # whitespace test
    body = r.get_json()
    # Should fail at duplicate_check or plan_params lookup, NOT at validation.
    assert body["status"] == "error"
    assert body.get("stage") != "validation", f"unexpected validation error: {body}"
    # plan_id should appear trimmed in the error payload
    assert body.get("plan_id") == "__DOES_NOT_EXIST__"


@check("flask: /generate- rejects non-string plan_id with 400")
def _():
    from app import create_app
    client = create_app().test_client()
    r = client.post("/app/v1/jkt/planning-scheduling/plan/generate-plan",
                    json={"plan_id": 12345})
    assert r.status_code == 400, r.status_code
    assert "must be a string" in r.get_json()["message"]


@check("flask: /generate- rejects malformed JSON with 400")
def _():
    from app import create_app
    client = create_app().test_client()
    r = client.post("/app/v1/jkt/planning-scheduling/plan/generate-plan",
                    data="not json at all",
                    content_type="application/json")
    assert r.status_code == 400, r.status_code
    assert r.get_json()["status"] == "error"


@check("flask: /generate- rejects already-scheduled plan_id with 409")
def _():
    """Find any plan_id with rows in jkt_plan_kpis; skip cleanly if there's none."""
    from V1.utilities import config_loader, db
    from app import create_app
    cfg = config_loader.load()
    conn = db.connect(cfg["db"])
    cur = conn.cursor()
    cur.execute("SELECT plan_id FROM jkt_plan_kpis LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        print("        (no scheduled plan_id in DB — skipping)")
        return
    scheduled_pid = row[0]
    client = create_app().test_client()
    r = client.post("/app/v1/jkt/planning-scheduling/plan/generate-plan",
                    json={"plan_id": scheduled_pid})
    assert r.status_code == 409, f"expected 409, got {r.status_code}: {r.get_json()}"
    body = r.get_json()
    assert body["status"] == "error"
    assert body["stage"] == "duplicate_check"


@check("exceptions: PipelineError class is importable")
def _():
    from V1.utilities.exceptions import PipelineError
    e = PipelineError("test", stage="x", status_code=418)
    assert e.stage == "x"
    assert e.status_code == 418
    assert str(e) == "test"


# --------------------------------------------------------------------------- #
# 8. Reports / writers (import-only, don't actually insert)                   #
# --------------------------------------------------------------------------- #
@check("reports: writer modules expose upload()")
def _():
    from V1.reports import kpi_writer, plan_writer, capacity_writer
    for m in (kpi_writer, plan_writer, capacity_writer):
        assert callable(m.upload), f"{m.__name__}.upload missing"


@check("kpi_writer: _count_demand_skus returns positive count")
def _():
    if _skip_if_no_test_plan("kpi_writer: _count_demand_skus returns positive count"): return
    from V1.utilities import config_loader
    from V1.reports.kpi_writer import _count_demand_skus
    cfg = config_loader.load()
    n = _count_demand_skus(TEST_PLAN_ID, cfg["db"])
    assert n > 0, f"expected positive distinct demand SKU count, got {n}"


@check("capacity_writer: imports plan_params (for window filtering)")
def _():
    """Verify the new dependency wiring is intact."""
    import V1.reports.capacity_writer as cw
    assert hasattr(cw, "plan_params"), "capacity_writer should import plan_params"


@check("schedule_route: _DEFAULT_CT_MIN constant + _post_process_schedule_excel function exist")
def _():
    """Verify the new cycle-time default + Excel post-process are wired in."""
    from V1.routes import schedule_route as sr
    assert hasattr(sr, "_DEFAULT_CT_MIN"), "missing _DEFAULT_CT_MIN constant"
    assert sr._DEFAULT_CT_MIN == 15.0, f"default CT should be 15, got {sr._DEFAULT_CT_MIN}"
    assert callable(getattr(sr, "_post_process_schedule_excel", None)), \
        "missing _post_process_schedule_excel(path, default_ct_skus)"


@check("schedule_route: default-CT SKUs run through (raw+buffer)/efficiency, not flat 15")
def _():
    """The 15-min default must be treated as a RAW cure time and pushed through
    the same (raw + buffer) / efficiency formula as every other SKU — NOT
    injected as a flat final cycle time."""
    import inspect
    from V1.routes import schedule_route as sr
    src = inspect.getsource(sr.run_from_database)
    assert "_DEFAULT_CT_MIN" in src, "default injection should reference _DEFAULT_CT_MIN"
    assert "LOAD_UNLOAD_BUFFER_MIN" in src and "PRESS_EFFICIENCY" in src, \
        "default cycle time must apply buffer + efficiency like every other SKU"
    # The old flat assignment {"CycleTime_min": _DEFAULT_CT_MIN} must be gone.
    assert 'CycleTime_min": _DEFAULT_CT_MIN' not in src, \
        "default CT should no longer be assigned flat — must go through the formula"
    # infeasibility_writer should report the EFFECTIVE CT from the Shift Schedule.
    iw_src = inspect.getsource(__import__("V1.reports.infeasibility_writer",
                                          fromlist=["_effective_cycle_times"]))
    assert "_effective_cycle_times" in iw_src and "Shift Schedule" in iw_src, \
        "infeasibility_writer must recover the effective CT from the Shift Schedule"


@check("schedule_route: post-process Excel excludes CO+clean AND sets 'NA' for default CT SKUs")
def _():
    """Build a tiny synthetic schedule Excel and run the post-process on it."""
    import openpyxl, tempfile, os
    from datetime import datetime
    from V1.routes.schedule_route import _post_process_schedule_excel
    wb = openpyxl.Workbook()
    # Demand Fulfillment sheet
    ws = wb.active; ws.title = "Demand Fulfillment"
    ws.append(["title"]); ws.append(["summary"])
    ws.append(["SKUCode","Priority","Demand","GT","Planned_Units","Gap","Fulfillment_Pct","Status","CycleTime_min"])
    ws.append(["SKU_REAL",   0.5, 100, 0, 100, 0, 1.0, "FULLY MET", 17])
    ws.append(["SKU_DEFAULT",0.5, 100, 0, 100, 0, 1.0, "FULLY MET", 15])
    # Shift Schedule
    ws2 = wb.create_sheet("Shift Schedule")
    ws2.append(["title"]); ws2.append(["summary"])
    ws2.append(["Date","Shift","Machine","SKUCode","StartTime","EndTime","Qty","CT","GT","Remarks"])
    ws2.append([datetime(2026,6,1), "A", "1234", "SKU_REAL",   datetime(2026,6,1,7), datetime(2026,6,1,15), 30, 16, 0, "LP Scheduled"])
    ws2.append([datetime(2026,6,1), "B", "1234", "CHANGEOVER", datetime(2026,6,1,15),datetime(2026,6,1,20), 0,  0, 0, "C/O"])
    ws2.append([datetime(2026,6,1), "B", "1234", "SKU_REAL",   datetime(2026,6,1,20),datetime(2026,6,1,22), 10, 12, 0, "Mould Cleaning"])
    # Machine Utilization
    ws3 = wb.create_sheet("Machine Utilization")
    ws3.append(["title"]); ws3.append(["summary"])
    ws3.append(["Machine","Avail","Used","Idle","Util","SKUs","Cycles","Units"])
    ws3.append([1234, 1440, 900, 540, 0.625, 1, 25, 50])
    # Machine Schedule (post-process will rewrite this)
    ws4 = wb.create_sheet("Machine Schedule")
    ws4.append(["title"]); ws4.append(["summary"])
    ws4.append(["Machine","SKUCode","Priority","CycleTime_min","Cycles","Units_Planned","Mins_Used","Days_Used"])
    ws4.append([1234, "SKU_REAL", 0.5, 17, 5, 10, 80, 0.06])  # legacy stub; will be rebuilt

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); f.close()
    wb.save(f.name); wb.close()
    _post_process_schedule_excel(f.name, {"SKU_DEFAULT"})

    wb2 = openpyxl.load_workbook(f.name, data_only=True)
    # Check NA was applied
    assert wb2["Demand Fulfillment"].cell(row=4, column=9).value == 17,   "real SKU's CT should stay 17"
    assert wb2["Demand Fulfillment"].cell(row=5, column=9).value == "NA","defaulted SKU should show NA"
    # Check Machine Utilization Used_Mins = 480 production only (excluded 300 CO + 120 clean)
    assert wb2["Machine Utilization"].cell(row=4, column=3).value == 480, \
        f"Used should be 480 (productive only), got {wb2['Machine Utilization'].cell(row=4, column=3).value}"
    # Machine Schedule was rebuilt: only the productive row counts (Mould Cleaning row excluded)
    ws_m = wb2["Machine Schedule"]
    found = False
    for r in range(4, ws_m.max_row + 1):
        m_val = ws_m.cell(row=r, column=1).value
        if str(m_val) == "1234" and ws_m.cell(row=r, column=2).value == "SKU_REAL":
            assert ws_m.cell(row=r, column=6).value == 30, \
                f"Units_Planned should = 30 (only productive row; clean+CO excluded), got {ws_m.cell(row=r, column=6).value}"
            found = True
    assert found, "rebuilt Machine Schedule row missing for (1234, SKU_REAL)"
    wb2.close(); os.unlink(f.name)


@check("schedule_route: Machine Schedule reconciles with Shift Schedule across multi-run SKUs")
def _():
    """Three productive runs on the same machine for the same SKU should
    aggregate to one row totaling 80 units (legacy LP would have shown only
    the first-pass 30; rebuild fixes this)."""
    import openpyxl, tempfile, os
    from datetime import datetime
    from V1.routes.schedule_route import _post_process_schedule_excel
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Demand Fulfillment"
    ws.append(["t"]); ws.append(["s"])
    ws.append(["SKUCode","Priority","Demand","GT","Planned_Units","Gap","Fulfillment_Pct","Status","CycleTime_min"])
    ws.append(["SKU_A", 0.9, 80, 0, 80, 0, 1.0, "FULLY MET", 18])
    ws2 = wb.create_sheet("Shift Schedule")
    ws2.append(["t"]); ws2.append(["s"])
    ws2.append(["Date","Shift","Machine","SKUCode","StartTime","EndTime","Qty","CT","GT","Remarks"])
    ws2.append([datetime(2026,6,1),"A","5555","SKU_A",datetime(2026,6,1,7),datetime(2026,6,1,12),30,18,0,"LP"])
    ws2.append([datetime(2026,6,2),"A","5555","SKU_A",datetime(2026,6,2,7),datetime(2026,6,2,12),30,18,0,"continuity"])
    ws2.append([datetime(2026,6,3),"A","5555","SKU_A",datetime(2026,6,3,7),datetime(2026,6,3,9),20,18,0,"extra run"])
    ws3 = wb.create_sheet("Machine Utilization")
    ws3.append(["t"]); ws3.append(["s"])
    ws3.append(["Machine","Avail","Used","Idle","Util","SKUs","Cycles","Units"])
    ws3.append([5555, 1440, 100, 1340, 0.07, 1, 50, 100])
    ws4 = wb.create_sheet("Machine Schedule")
    ws4.append(["t"]); ws4.append(["s"])
    ws4.append(["Machine","SKUCode","Priority","CycleTime_min","Cycles","Units_Planned","Mins_Used","Days_Used"])
    ws4.append([5555, "SKU_A", 0.9, 18, 15, 30, 270, 0.19])  # buggy first-pass: only 30

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); f.close()
    wb.save(f.name); wb.close()
    _post_process_schedule_excel(f.name, set())

    wb2 = openpyxl.load_workbook(f.name, data_only=True)
    ws_m = wb2["Machine Schedule"]
    total = sum(float(ws_m.cell(row=r, column=6).value or 0) for r in range(4, ws_m.max_row+1))
    assert total == 80, f"Σ Units_Planned should = 80 (multi-run aggregated), got {total}"
    wb2.close(); os.unlink(f.name)


@check("schedule_route: DB overrides for noOfChangeOver + efficiency are wired in")
def _():
    """Verify the new DB-override block exists in _run_locked()."""
    src = __import__("pathlib").Path("V1/routes/schedule_route.py").read_text()
    assert "noOfChangeOver" in src, "noOfChangeOver override missing from schedule_route"
    assert "MAX_CHANGEOVERS_PER_SHIFT = int(db_co_per_shift)" in src, \
        "MAX_CHANGEOVERS_PER_SHIFT should be set DIRECTLY from DB (per-shift, no /3 division)"
    assert "PRESS_EFFICIENCY = float(db_eff) / 100" in src, "efficiency not overridden from DB"
    # Ensure the OLD per-day division logic is gone
    assert "Config.SHIFTS_PER_DAY" not in (
        src.split("noOfChangeOver", 1)[1].split("efficiency", 1)[0] if "noOfChangeOver" in src else ""
    ), "Old /SHIFTS_PER_DAY division still present in noOfChangeOver block"


@check("schedule_route: run() is serialized with a lock")
def _():
    """Confirm the threading.Lock is in place and run() acquires it."""
    import V1.routes.schedule_route as sr
    import threading
    assert hasattr(sr, "_RUN_LOCK"), "expected _RUN_LOCK at module level"
    assert isinstance(sr._RUN_LOCK, type(threading.Lock())), "_RUN_LOCK must be a Lock"
    assert hasattr(sr, "_run_locked"), "expected _run_locked inner function"


@check("plan_writer: _load_sku_descriptions returns SKU→desc map")
def _():
    if _skip_if_no_test_plan("plan_writer: _load_sku_descriptions returns SKU→desc map"): return
    from V1.utilities import config_loader
    from V1.reports.plan_writer import _load_sku_descriptions
    cfg = config_loader.load()
    mp = _load_sku_descriptions(TEST_PLAN_ID, cfg["db"])
    assert isinstance(mp, dict) and len(mp) > 0, f"expected non-empty map, got {mp!r}"
    for sku, desc in mp.items():
        assert isinstance(desc, str) and desc.strip(), f"bad desc for {sku!r}: {desc!r}"


@check("plan_writer: _round_sku_totals_up_to_even bumps odd SKUs by +1")
def _():
    from V1.reports.plan_writer import _round_sku_totals_up_to_even
    # Two SKUs, A=odd total (101), B=even total (100). CHANGEOVER row skipped.
    rows = [
        ("pid", "A", "desc", None, None, None, None, 50, None, None, None, "by"),
        ("pid", "A", "desc", None, None, None, None, 51, None, None, None, "by"),   # total A = 101 odd
        ("pid", "B", "desc", None, None, None, None, 60, None, None, None, "by"),
        ("pid", "B", "desc", None, None, None, None, 40, None, None, None, "by"),   # total B = 100 even
        ("pid", "CHANGEOVER", None, None, None, None, None, 0, None, None, None, "by"),
    ]
    bumped = _round_sku_totals_up_to_even(rows)
    assert bumped == 1, f"expected 1 SKU bumped, got {bumped}"
    a_total = sum(r[7] for r in rows if r[1] == "A")
    b_total = sum(r[7] for r in rows if r[1] == "B")
    assert a_total == 102 and a_total % 2 == 0, f"A total should be even 102, got {a_total}"
    assert b_total == 100, f"B unchanged, got {b_total}"


@check("kpi_writer: demand-weighted fulfillment rounds planned UP to even per SKU")
def _():
    import openpyxl
    from V1.reports.kpi_writer import _demand_weighted_fulfillment
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["title"]); ws.append(["summary"])
    ws.append(["SKUCode","Priority","Demand","GT","Planned_Units","Gap","Fulfillment_Pct"])
    # Two SKUs, equal demand 100. Planned 99 (odd→100) vs 100. Expected = 100% capped.
    ws.append(["A", 0, 100, 0,  99, 0, 0.99])
    ws.append(["B", 0, 100, 0, 100, 0, 1.00])
    result = _demand_weighted_fulfillment(ws)
    assert abs(result - 100.0) < 0.01, f"expected 100.0 (both round to 100% capped), got {result}"


@check("kpi_writer: _count_planned_skus excludes TOTAL row AND skus with Planned_Units=0")
def _():
    """planSKU should count only SKUs that actually got production:
       - skip 'TOTAL' summary row
       - skip SKUs with Planned_Units == 0 (status UNMET / UNSCHEDULABLE)"""
    import openpyxl
    from V1.reports.kpi_writer import _count_planned_skus, _demand_weighted_fulfillment
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["title"]); ws.append(["summary"])
    ws.append(["SKUCode","Priority","Demand","GT","Planned_Units","Gap","Fulfillment_Pct"])
    ws.append(["A",      0, 100, 0, 100, 0, 1.0])      # planned > 0  → count
    ws.append(["B",      0, 100, 0,  80, 0, 0.8])      # planned > 0  → count
    ws.append(["C_UNMET",0, 100, 0,   0, 100, 0.0])    # planned = 0  → DO NOT count
    ws.append(["TOTAL",  0, 300, 0, 180, 120, 0.6])    # summary row  → DO NOT count
    n = _count_planned_skus(ws)
    assert n == 2, f"expected planSKU=2 (only A and B have Planned>0), got {n}"
    # Fulfillment should also exclude the TOTAL row but include C (it has demand>0)
    f = _demand_weighted_fulfillment(ws)
    # (1.0·100 + 0.8·100 + 0.0·100) / 300 = 60%
    assert abs(f - 60.0) < 0.01, f"expected 60.0%, got {f}"


@check("kpi_writer: demand-weighted fulfillment caps each SKU at 100%")
def _():
    """Synthetic sheet: SKU A over-fulfilled (120%), SKU B at 80%, equal demand.
    Capped weighted = (1.0·100 + 0.8·100)/200 = 90%. Uncapped would be 100%."""
    import openpyxl
    from V1.reports.kpi_writer import _demand_weighted_fulfillment
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["title"]); ws.append(["summary"])
    ws.append(["SKUCode","Priority","Demand","GT","Planned_Units","Gap","Fulfillment_Pct"])
    ws.append(["A", 0, 100, 0, 120, 0, 1.2])   # over-fulfilled → capped to 1.0
    ws.append(["B", 0, 100, 0,  80, 0, 0.8])
    result = _demand_weighted_fulfillment(ws)
    assert abs(result - 90.0) < 0.01, f"expected 90.0, got {result}"


@check("capacity_writer: skips CHANGEOVER + mould-clean rows from busy time")
def _():
    """Productive util only — CO and clean rows must NOT count as busy."""
    import openpyxl
    from datetime import datetime, date
    from V1.reports.capacity_writer import compute_daily_utilisation
    # Build a minimal workbook with 1 machine, 1 day, and a mix of rows
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Shift Schedule"
    ws.append(["title"]); ws.append(["summary"])
    ws.append(["Date","Shift","Machine","SKUCode","StartTime","EndTime","Qty","CycleTime_min","GT","Remarks"])
    d_obj = datetime(2026, 6, 1, 7, 0)
    # 480 min production + 300 min CHANGEOVER + 120 min CLEAN  =  900 min total, 480 productive
    ws.append([d_obj, "A", "M1", "SKU_A",     d_obj,                  datetime(2026,6,1,15,0), 30, 16, 0, "LP Scheduled"])
    ws.append([d_obj, "B", "M1", "CHANGEOVER",datetime(2026,6,1,15,0),datetime(2026,6,1,20,0),  0,  0, 0, "C/O"])
    ws.append([d_obj, "B", "M1", "SKU_A",     datetime(2026,6,1,20,0),datetime(2026,6,1,22,0), 10, 12, 0, "Mould Cleaning"])
    # Machine Utilization sheet — fleet of 1
    ws_u = wb.create_sheet("Machine Utilization")
    ws_u.append(["title"]); ws_u.append(["summary"])
    ws_u.append(["Machine","Avail","Used","Idle","Util","SKUs","Cycles","Units"])
    ws_u.append(["M1", 1440, 480, 960, 0.33, 1, 0, 60])

    daily = compute_daily_utilisation(wb, date(2026,6,1), date(2026,6,1))
    util = daily[0][1]
    # 480 productive min / 1440 = 33.33% (NOT 900/1440 = 62.5%)
    assert abs(util - 33.33) < 0.5, f"expected ~33.33% (productive only), got {util}"


@check("simulation: package importable + Flask blueprint registered")
def _():
    """The simulation/ folder defines a Blueprint with prefix /app/v1/jkt/planning-scheduling
    and route /simulation/generate-plan. Verify both load cleanly."""
    import importlib
    sim_api = importlib.import_module("simulation.routes.api_route")
    assert hasattr(sim_api, "bp"), "simulation.routes.api_route must export bp"
    assert sim_api.bp.name == "simulation", f"blueprint name should be 'simulation', got {sim_api.bp.name!r}"
    assert sim_api.API_SIMULATE_PATH == "/simulation/generate-plan"
    # Thin route wrappers re-export V1's run() — confirm
    from simulation.routes import demand_route as sd, schedule_route as ss, upload_route as su
    from V1.routes import demand_route as vd, schedule_route as vs, upload_route as vu
    assert sd.run is vd.run, "simulation demand_route should re-export V1's run"
    assert ss.run is vs.run, "simulation schedule_route should re-export V1's run"
    assert su.run is vu.run, "simulation upload_route should re-export V1's run"


@check("simulation: apply_mode('simulation') resolves to jkt_sim_* tables")
def _():
    """The mode_token system in config_loader inserts 'sim_' after 'jkt_'."""
    from V1.utilities import config_loader
    cfg = config_loader.load(mode="simulation")
    assert cfg["mode"] == "simulation"
    tbl = cfg["tbl"]
    # All sim tables must include 'jkt_sim_' prefix
    expected = {
        "demand":        "jkt_sim_demand",
        "plan_params":   "jkt_sim_plan_params",
        "plan_kpis":     "jkt_sim_plan_kpis",
        "plan":          "jkt_sim_plan",
        "capacity":      "jkt_sim_plan_capacityUtilisation",
    }
    for logical, physical in expected.items():
        assert tbl.get(logical) == physical, \
            f"sim mode: tbl[{logical!r}] should be {physical!r}, got {tbl.get(logical)!r}"


@check("simulation: planning mode still resolves to jkt_* tables (no regression)")
def _():
    """Ensure removing /simulation from V1's blueprint didn't break planning mode."""
    from V1.utilities import config_loader
    cfg = config_loader.load(mode="planning")
    assert cfg["mode"] == "planning"
    assert cfg["tbl"]["demand"] == "jkt_demand"
    assert cfg["tbl"]["plan_params"] == "jkt_plan_params"
    assert cfg["tbl"]["plan_kpis"] == "jkt_plan_kpis"
    # V1's api_route blueprint should still expose /plan/generate-plan
    from V1.routes import api_route as v1_api
    assert v1_api.API_GENERATE_PATH == "/plan/generate-plan"
    # And should NOT have a simulation route anymore (moved to simulation/)
    assert not hasattr(v1_api, "API_SIMULATE_PATH"), \
        "V1 api_route should no longer expose API_SIMULATE_PATH (moved to simulation/routes/api_route.py)"


@check("simulation: sim_status delegates to V1's plan_status with ALL 4 sim table names")
def _():
    """sim_status.assert_not_already_simulated() must pass ALL FOUR sim output
    table names (kpis, plan, capacity, infeasibility) to plan_status. The
    Infeasibility table has an auto-increment PK with NO DB-level dedupe; if
    the duplicate check omits it, a re-run after partial cleanup silently
    appends duplicate infeasibility rows."""
    from simulation.setups import sim_status
    import inspect
    src = inspect.getsource(sim_status.assert_not_already_simulated)
    assert "plan_status.assert_not_already_scheduled" in src, \
        "sim_status should delegate to V1's plan_status"
    for key in ("plan_kpis", "plan", "capacity", "infeasibility"):
        assert f'tbl["{key}"]' in src, \
            f"sim_status duplicate-check must include tbl[{key!r}] " \
            f"(otherwise re-runs silently duplicate {key} rows)"


@check("planning: api_route duplicate-check includes ALL 4 output tables (incl. infeasibility)")
def _():
    """Mirror of the sim check — the planning path must ALSO pass the
    infeasibility table to plan_status, since the writer is called in
    Phase C for BOTH modes."""
    import inspect
    from V1.routes import api_route as v1_api
    src = inspect.getsource(v1_api._generate)
    for key in ("plan_kpis", "plan", "capacity", "infeasibility"):
        assert f'cfg["tbl"]["{key}"]' in src, \
            f"V1 api_route._generate must include cfg['tbl'][{key!r}] in output_tables"


@check("simulation: app.py registers BOTH planning + simulation blueprints")
def _():
    src = __import__("pathlib").Path("app.py").read_text()
    assert "planning_bp" in src and "simulation_bp" in src, \
        "app.py must register both blueprints"
    assert "from simulation.routes.api_route import bp as simulation_bp" in src
    # Confirm the live app actually has both URL rules
    import importlib
    app_module = importlib.import_module("app")
    rules = {r.rule for r in app_module.app.url_map.iter_rules()}
    assert "/app/v1/jkt/planning-scheduling/plan/generate-plan" in rules, \
        f"planning route missing from URL map: {rules}"
    assert "/app/v1/jkt/planning-scheduling/simulation/generate-plan" in rules, \
        f"simulation route missing from URL map: {rules}"


@check("filenames: mode_file_tag is '' for planning, 'sim_' for simulation")
def _():
    """Planning Excels keep their exact filename; simulation gets a 'sim_' marker
    so the two pipelines never overwrite each other's output for the same plan_id."""
    from V1.utilities import config_loader
    plan_cfg = config_loader.load(mode="planning")
    sim_cfg  = config_loader.load(mode="simulation")
    assert config_loader.mode_file_tag(plan_cfg) == "", \
        f"planning tag should be empty, got {config_loader.mode_file_tag(plan_cfg)!r}"
    assert config_loader.mode_file_tag(sim_cfg) == "sim_", \
        f"simulation tag should be 'sim_', got {config_loader.mode_file_tag(sim_cfg)!r}"


@check("filenames: both Excel templates carry {mode_tag} and resolve distinctly")
def _():
    """The demand + schedule filename templates must include {mode_tag}, and the
    two modes must resolve to DIFFERENT filenames for the same plan_id."""
    from V1.utilities import config_loader
    for mode in ("planning", "simulation"):
        cfg = config_loader.load(mode=mode)
        assert "{mode_tag}" in cfg["demand"]["output_excel"], \
            "demand.output_excel must contain {mode_tag}"
        assert "{mode_tag}" in cfg["schedule"]["output_excel"], \
            "schedule.output_excel must contain {mode_tag}"
    # Resolve the demand filename for both modes with the same plan_id.
    pcfg = config_loader.load(mode="planning");   pcfg["plan"]["plan_id"] = "PID1"
    scfg = config_loader.load(mode="simulation"); scfg["plan"]["plan_id"] = "PID1"
    p_name = config_loader.resolve_paths(pcfg)["demand"]["output_excel"]
    s_name = config_loader.resolve_paths(scfg)["demand"]["output_excel"]
    assert p_name == "requirement_summary_PID1.xlsx", p_name
    assert s_name == "requirement_summary_sim_PID1.xlsx", s_name
    assert p_name != s_name, "planning and simulation must not share a filename"
    # Schedule template resolves with mode_tag too (dates filled by the route).
    s_sched = scfg["schedule"]["output_excel"].format(
        mode_tag=config_loader.mode_file_tag(scfg),
        plan_id="PID1", plan_start="2026-01-01", planning_days=30,
    )
    assert s_sched.startswith("PCR_Schedule_sim_PID1_"), s_sched


@check("infeasibility_writer: importable + wired into upload_route")
def _():
    """The 4th output writer exists, exposes upload(), and Phase C calls it."""
    import importlib
    iw = importlib.import_module("V1.reports.infeasibility_writer")
    assert hasattr(iw, "upload"), "infeasibility_writer must export upload()"
    up_src = __import__("pathlib").Path("V1/routes/upload_route.py").read_text()
    assert "infeasibility_writer" in up_src and "infeasibility_writer.upload(" in up_src, \
        "upload_route must call infeasibility_writer.upload()"


@check("infeasibility_writer: captures unmet demand AND default-CT (NA) SKUs")
def _():
    """Source-level guarantees: creates table non-destructively, flags NA cycle
    time as default-15-used, reads the right Demand Fulfillment columns, and
    delegates row-filtering to the shared kpi_writer helpers."""
    import inspect
    from V1.reports import infeasibility_writer as iw
    up_src = inspect.getsource(iw.upload)
    # Default-CT / NA handling
    assert '"NA"' in up_src or "'NA'" in up_src, "must detect the 'NA' cycle-time marker"
    assert "defaultCycleTime" in inspect.getsource(iw), "must record a defaultCycleTime flag"
    assert "_DEFAULT_CT_MIN" in up_src, "must pull the default cycle time from schedule_route"
    # Non-destructive table creation (append-only invariant preserved)
    ensure_src = inspect.getsource(iw._ensure_table)
    assert "CREATE TABLE IF NOT EXISTS" in ensure_src, "must create table idempotently"
    assert "DROP" not in ensure_src and "ALTER" not in ensure_src, \
        "must never drop/alter an existing table"
    # Reuses shared helpers rather than re-implementing row filtering / coercion
    mod_src = inspect.getsource(iw)
    assert "_is_real_sku_row" in mod_src and "_safe_number" in mod_src, \
        "should reuse kpi_writer's _is_real_sku_row / _safe_number"


@check("infeasibility: apply_mode resolves jkt_plan_Infeasibility -> jkt_sim_plan_Infeasibility")
def _():
    """The 4th table name must be mode-aware exactly like the other outputs."""
    from V1.utilities import config_loader
    p = config_loader.load(mode="planning")["tbl"]["infeasibility"]
    s = config_loader.load(mode="simulation")["tbl"]["infeasibility"]
    assert p == "jkt_plan_Infeasibility", p
    assert s == "jkt_sim_plan_Infeasibility", s


@check("capacity_writer: divides by full fleet, not just used machines")
def _():
    """Read the source to confirm the denominator switched to all_machines."""
    src = (
        __import__("pathlib").Path("V1/reports/capacity_writer.py").read_text()
    )
    assert "all_machines" in src, "all_machines set should exist"
    assert "Machine Utilization" in src, "should read Machine Utilization sheet"
    assert "n_machines" in src, "should use fleet-size denominator"
    # Ensure the old machines_per_day denominator pattern is gone
    assert "machines_per_day[d]" not in src, "still using used-machines denominator"


# --------------------------------------------------------------------------- #
# Runner                                                                      #
# --------------------------------------------------------------------------- #
_SKIP_SENTINEL = object()


def main() -> int:
    if TEST_PLAN_ID is None:
        print("  (no plan_id with both demand and params found in DB — demand-dependent checks will skip)")
    else:
        print(f"  (using TEST_PLAN_ID = {TEST_PLAN_ID})")
    print()

    passed = failed = skipped = 0
    for name, fn in CHECKS:
        try:
            result = fn()
            # _skip_if_no_test_plan returns early (None) AFTER printing its own "SKIP" line.
            # The function body that follows is a no-op, so we can't easily tell pass vs skip
            # here — but the SKIP line already printed; just don't print PASS for those.
            if TEST_PLAN_ID is None and "TEST_PLAN_ID" in fn.__code__.co_names:
                skipped += 1
                continue
            print(f"  PASS  {name}")
            passed += 1
        except Exception as e:
            print(f"  FAIL  {name}")
            print(f"        {type(e).__name__}: {e}")
            traceback.print_exc(limit=2)
            failed += 1
    print(f"\n{passed} passed, {skipped} skipped, {failed} failed")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
