"""Insert one row into jkt_plan_kpis from the schedule's Demand Fulfillment summary."""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import openpyxl

from V1.reports.capacity_writer import compute_daily_utilisation
from V1.setups import plan_params
from V1.utilities.db import connect
from V1.utilities.time_utils import now_ist


def _find(pattern: str, text: str, default=None):
    m = re.search(pattern, text)
    if not m:
        return default
    return m.group(1)


def _require(pattern: str, text: str):
    v = _find(pattern, text)
    if v is None:
        raise ValueError(f"Pattern not found in summary: {pattern}")
    return v


def _is_real_sku_row(ws, r: int) -> bool:
    """Detail rows in the Demand Fulfillment sheet start at row 4. The legacy
    scheduler writes a 'TOTAL' summary row at the bottom — we MUST exclude it
    from per-SKU counts and aggregations, otherwise planSKU is off by +1 and
    the demand-weighted fulfillment double-counts demand.
    """
    sku = ws.cell(row=r, column=1).value
    if not sku:
        return False
    return str(sku).strip().upper() not in ("TOTAL", "GRAND TOTAL")


def _count_planned_skus(ws) -> int:
    """planSKU = count of SKUs that actually got production (Planned_Units > 0).

    Excludes:
      - the 'TOTAL' summary row at the bottom
      - SKUs with Planned_Units == 0 (status UNMET / UNSCHEDULABLE)

    Demand Fulfillment column 5 = Planned_Units.
    """
    n = 0
    for r in range(4, ws.max_row + 1):
        if not _is_real_sku_row(ws, r):
            continue
        planned = ws.cell(row=r, column=5).value or 0
        if planned > 0:
            n += 1
    return n


def _count_demand_skus(plan_id: str, db_cfg: dict) -> int:
    """Distinct SKUs requested in jkt_demand for this plan."""
    conn = connect(db_cfg)
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT COUNT(DISTINCT skuCode) FROM jkt_demand WHERE plan_id = %s",
            (plan_id,),
        )
        return int(cur.fetchone()[0])
    finally:
        cur.close()
        conn.close()


def _round_up_to_even(n: int) -> int:
    """Single source of truth for the +1-tyre rule. Used by BOTH kpi_writer
    and plan_writer so they can never disagree on what's 'rounded'."""
    n = int(n)
    return n + (n % 2)


def _safe_number(v, default: float = 0.0) -> float:
    """Robust cell-value → float. Tolerates messy Excel data: '#REF!', NaN,
    string-with-comma, None, etc. Returns `default` on any failure."""
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        f = float(v)
    else:
        try:
            f = float(str(v).strip().replace(",", ""))
        except (ValueError, TypeError):
            return default
    import math
    return f if math.isfinite(f) and f >= 0 else default


def _demand_weighted_fulfillment(ws) -> float:
    """Overall demand fulfillment as a demand-weighted average of per-SKU
    fulfillment, with each SKU capped at 100%:

        Σ_i [ min(planned_i / demand_i, 1.0) · (demand_i / Σ demand) ]  × 100

    Capping prevents over-produced SKUs from masking shortfalls on others.
    Demand Fulfillment sheet columns: 3 = Demand, 5 = Planned_Units.

    Plant constraint: per-SKU planned is rounded UP to the next even number
    (via shared `_round_up_to_even` helper — guarantees same value as plan_writer).
    """
    total_demand = 0.0
    weighted = 0.0
    for r in range(4, ws.max_row + 1):
        if not _is_real_sku_row(ws, r):
            continue                                # skip 'TOTAL' summary row
        demand  = _safe_number(ws.cell(row=r, column=3).value)
        planned = _safe_number(ws.cell(row=r, column=5).value)
        if demand <= 0:
            continue
        planned = _round_up_to_even(planned)        # shared rule with plan_writer
        total_demand += demand
        weighted += min(planned / demand, 1.0) * demand
    if total_demand == 0:
        print("[upload:kpi] WARNING — total_demand=0 in Demand Fulfillment sheet, "
              "demandFulfillment will be reported as 0%")
        return 0.0
    return round(weighted / total_demand * 100, 2)


def _overall_capacity_utilisation(wb, plan_id: str, db_cfg: dict) -> float:
    """Mean of the per-date fleet utilisations — same full-day (1440 min) math
    the capacity_writer uses, so the KPI matches jkt_plan_capacityUtilisation."""
    plan_row = plan_params.fetch(db_cfg, plan_id)
    ps, pe = plan_row["planStartDate"], plan_row["planEndDate"]
    if isinstance(ps, datetime): ps = ps.date()
    if isinstance(pe, datetime): pe = pe.date()
    daily = compute_daily_utilisation(wb, ps, pe)
    if not daily:
        return 0.0
    return round(sum(u for _, u in daily) / len(daily), 2)


def upload(schedule_path: Path, plan_id: str, created_by: str, db_cfg: dict) -> None:
    wb = openpyxl.load_workbook(schedule_path, data_only=True)
    try:
        ws = wb["Demand Fulfillment"]
        summary = ws.cell(row=2, column=1).value or ""

        demand_sku = _count_demand_skus(plan_id, db_cfg)
        plan_sku   = _count_planned_skus(ws)

        row = {
            "plan_id":             plan_id,
            "demandFulfillment":   _demand_weighted_fulfillment(ws),
            "demandSKU":           demand_sku,
            "planSKU":             plan_sku,
            "capacityUtilisation": _overall_capacity_utilisation(wb, plan_id, db_cfg),
            "curingChangeovers":   int(_require(r"Changeovers:\s*([\d,]+)", summary).replace(",", "")),
            "createdAt":           now_ist(),
            "createdBy":           created_by,
        }
    finally:
        wb.close()                                       # release file handle

    conn = connect(db_cfg)
    try:
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO jkt_plan_kpis
                   (plan_id, demandFulfillment, demandSKU, planSKU,
                    capacityUtilisation, curingChangeovers, createdAt, createdBy)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
            (row["plan_id"], row["demandFulfillment"], row["demandSKU"], row["planSKU"],
             row["capacityUtilisation"], row["curingChangeovers"], row["createdAt"], row["createdBy"]),
        )
        conn.commit()
        print(f"[upload:kpi] inserted 1 row into jkt_plan_kpis")
    finally:
        cur.close()
        conn.close()
