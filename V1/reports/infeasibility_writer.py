"""Insert one row per at-risk SKU into jkt_plan_Infeasibility.

This is the 4th output table (reserved name `infeasibility` in config.yaml).
It records, for a given plan_id, every SKU whose demand could NOT be fully &
cleanly met by the LP schedule — so planners can see WHAT fell short and WHY
without opening the Excel. Two categories are captured:

  1. Demand shortfalls — Status in {PARTIAL, UNMET, UNSCHEDULABLE} (Gap > 0).
     The LP's own `Skip_Reason` (no eligible machine, no mould, capacity, …)
     is carried through verbatim.

  2. Default-cycle-time SKUs — rows whose `CycleTime_min` shows "NA" in the
     Demand Fulfillment sheet. These SKUs had no row in
     Master_Curing_Design_CycleTime, so the scheduler assumed a DEFAULT raw
     cure of 15 min (see schedule_route._DEFAULT_CT_MIN), run through the same
     (raw + buffer) / efficiency formula as every other SKU. Their plan is
     therefore based on an *assumed* cycle time and should be reviewed — hence
     flagged here even when demand was fully met. The `cycleTime` reported is
     the EFFECTIVE value the LP used (recovered from the Shift Schedule).

A SKU can fall into BOTH categories; `reason` then mentions both.

Source columns (Demand Fulfillment sheet, 1-indexed; data starts at row 4):
    1 SKUCode  3 Demand  5 Planned_Units  6 Gap  7 Fulfillment_Pct
    8 Status   9 CycleTime_min ("NA" => default used)   12 Skip_Reason

Mode-aware: the physical table name comes from cfg["tbl"]["infeasibility"]
  planning   -> jkt_plan_Infeasibility
  simulation -> jkt_sim_plan_Infeasibility

Schema (created on first run via CREATE TABLE IF NOT EXISTS — non-destructive):
    id BIGINT AUTO_INCREMENT PK, plan_id, skuCode, skuDescription,
    demand, plannedUnits, gap, fulfillmentPct, status, reason,
    cycleTime, defaultCycleTime (1=NA/default-15-used), createdAt, createdBy
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

# Reuse the SAME sheet-parsing helpers the KPI writer uses, so the two writers
# can never disagree on which rows are real SKUs / how cells are coerced.
from V1.reports.kpi_writer import _is_real_sku_row, _round_up_to_even, _safe_number
from V1.utilities.db import connect, safe_table
from V1.utilities.time_utils import now_ist

# Demand Fulfillment column indices (1-based).
_COL_SKU      = 1
_COL_DEMAND   = 3
_COL_PLANNED  = 5
_COL_GAP      = 6
_COL_FULFIL   = 7
_COL_STATUS   = 8
_COL_CT       = 9
_COL_SKIP     = 12

_BATCH = 1000


def _load_sku_descriptions(
    plan_id: str, db_cfg: dict, demand_table: str = "jkt_demand"
) -> dict[str, str]:
    """SKUCode -> description from the demand table (same source plan_writer uses)."""
    demand_table = safe_table(demand_table)
    conn = connect(db_cfg)
    try:
        cur = conn.cursor()
        cur.execute(
            f"SELECT skuCode, MAX(skuDescription) FROM {demand_table} "
            "WHERE plan_id = %s GROUP BY skuCode",
            (plan_id,),
        )
        return {sku: desc for sku, desc in cur.fetchall() if sku}
    finally:
        cur.close()
        conn.close()


def _ensure_table(cur, table: str) -> None:
    """Create the infeasibility table if it doesn't exist yet.

    Idempotent + non-destructive (IF NOT EXISTS), so it respects the pipeline's
    append-only invariant — it never alters or drops an existing table. The name
    is already validated by safe_table() at the call site.
    """
    cur.execute(
        f"""CREATE TABLE IF NOT EXISTS {table} (
                id              BIGINT       NOT NULL AUTO_INCREMENT,
                plan_id         VARCHAR(50)  NOT NULL,
                skuCode         VARCHAR(100),
                skuDescription  VARCHAR(255),
                demand          INT,
                plannedUnits    INT,
                gap             INT,
                fulfillmentPct  DECIMAL(6,2),
                status          VARCHAR(30),
                reason          VARCHAR(500),
                cycleTime       DECIMAL(8,2),
                defaultCycleTime TINYINT(1)  NOT NULL DEFAULT 0,
                createdAt       DATETIME,
                createdBy       VARCHAR(100),
                PRIMARY KEY (id),
                KEY idx_plan_id (plan_id)
            )"""
    )


def _build_reason(status: str, gap: int, skip_reason: str, default_ct: bool,
                  default_ct_min: float) -> str:
    """Human-readable explanation, combining the LP's skip reason with the
    default-cycle-time note when applicable."""
    parts: list[str] = []
    if skip_reason:
        parts.append(str(skip_reason).strip())
    if not skip_reason and gap > 0:
        parts.append(f"Demand shortfall of {gap} unit(s) ({status or 'not fully met'})")
    if default_ct:
        parts.append(
            f"Cycle time missing in master — assumed {default_ct_min:g} min cure "
            f"(buffer + efficiency applied)"
        )
    return "; ".join(p for p in parts if p) or (status or "infeasible")


def _effective_cycle_times(wb) -> dict[str, float]:
    """SKUCode -> the cycle time the scheduler ACTUALLY used, read from the
    Shift Schedule sheet (col 4 = SKUCode, col 8 = CycleTime_min).

    For default-CT SKUs the Demand Fulfillment sheet shows "NA", so this is how
    we recover the real number the LP used — the default 15 min cure run through
    (raw + buffer) / efficiency. SKUs that were never scheduled (UNSCHEDULABLE)
    don't appear here and are reported with a NULL cycleTime.
    """
    out: dict[str, float] = {}
    if "Shift Schedule" not in wb.sheetnames:
        return out
    ss = wb["Shift Schedule"]
    for r in range(4, ss.max_row + 1):
        sku = ss.cell(row=r, column=4).value
        ct  = ss.cell(row=r, column=8).value
        if sku and str(sku) not in out and ct not in (None, ""):
            val = _safe_number(ct)
            if val > 0:
                out[str(sku)] = val
    return out


def upload(
    schedule_path: Path, plan_id: str, created_by: str, db_cfg: dict,
    tables: dict | None = None,
) -> None:
    tables       = tables or {}
    demand_table = tables.get("demand", "jkt_demand")
    infeas_table = safe_table(tables.get("infeasibility", "jkt_plan_Infeasibility"))

    # Default cycle time the scheduler assumes for SKUs missing in the CT master.
    # Lazily imported to keep this module light to import (schedule_route is heavy).
    from V1.routes.schedule_route import _DEFAULT_CT_MIN as DEFAULT_CT

    sku_desc = _load_sku_descriptions(plan_id, db_cfg, demand_table)

    wb = openpyxl.load_workbook(schedule_path, data_only=True)
    try:
        ws = wb["Demand Fulfillment"]
        sched_ct = _effective_cycle_times(wb)   # SKU -> CT actually used by the LP
        now = now_ist()
        rows: list[tuple] = []
        for r in range(4, ws.max_row + 1):
            if not _is_real_sku_row(ws, r):
                continue                                  # skip TOTAL row
            sku     = ws.cell(row=r, column=_COL_SKU).value
            demand  = int(_safe_number(ws.cell(row=r, column=_COL_DEMAND).value))
            planned = int(_safe_number(ws.cell(row=r, column=_COL_PLANNED).value))
            gap     = int(_safe_number(ws.cell(row=r, column=_COL_GAP).value))
            fulfil  = round(_safe_number(ws.cell(row=r, column=_COL_FULFIL).value), 2)
            status  = (ws.cell(row=r, column=_COL_STATUS).value or "")
            status  = str(status).strip()
            ct_raw  = ws.cell(row=r, column=_COL_CT).value
            skip    = ws.cell(row=r, column=_COL_SKIP).value or ""

            # "NA" in the CycleTime_min column means the SKU had no master CT and
            # the LP fell back to the default (post-process writes the literal "NA").
            default_ct = str(ct_raw).strip().upper() == "NA"

            # Demand is "not fully met" when there's a positive gap OR the status
            # is anything other than FULLY MET. Use even-rounded planned (the same
            # rule kpi_writer/plan_writer apply) so this agrees with the KPIs.
            gap_is_real = _round_up_to_even(planned) < demand
            infeasible  = gap_is_real or (status.upper() not in ("FULLY MET", ""))

            if not (infeasible or default_ct):
                continue                                  # SKU is fine — skip

            # Report gap based on the even-rounded planned (never negative).
            report_gap = max(0, demand - _round_up_to_even(planned))
            # For default-CT SKUs the sheet shows "NA", so report the EFFECTIVE
            # cycle time the LP used (default cure run through buffer + efficiency),
            # recovered from the Shift Schedule; None if the SKU was never scheduled.
            if default_ct:
                cycle_time = sched_ct.get(str(sku))
            elif ct_raw not in (None, ""):
                cycle_time = _safe_number(ct_raw)
            else:
                cycle_time = None
            rows.append((
                plan_id,
                sku,
                sku_desc.get(sku),
                demand,
                planned,
                report_gap,
                fulfil,
                status or None,
                _build_reason(status, report_gap, skip, default_ct, DEFAULT_CT),
                cycle_time,
                1 if default_ct else 0,
                now,
                created_by,
            ))
    finally:
        wb.close()

    conn = connect(db_cfg)
    try:
        cur = conn.cursor()
        _ensure_table(cur, infeas_table)
        total = 0
        for i in range(0, len(rows), _BATCH):
            cur.executemany(
                f"""INSERT INTO {infeas_table}
                       (plan_id, skuCode, skuDescription, demand, plannedUnits,
                        gap, fulfillmentPct, status, reason, cycleTime,
                        defaultCycleTime, createdAt, createdBy)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                rows[i:i + _BATCH],
            )
            total += cur.rowcount
        conn.commit()
        print(f"[upload:infeasibility] inserted {total} rows into {infeas_table}")
    finally:
        cur.close()
        conn.close()
