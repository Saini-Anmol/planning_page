"""Compute date-wise capacity utilisation from the schedule's Shift Schedule sheet
and bulk-insert into jkt_plan_capacityUtilisation.

Method: used_minutes = EndTime − StartTime per slot, split at midnight so each
date only gets its own minutes. Per-machine cap at 1440 (100%). Date-level
utilisation = average across machines that appear on that date. Stored as
percentage (e.g. 99.41).
"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from V1.setups import plan_params
from V1.utilities.db import connect
from V1.utilities.exceptions import PipelineError
from V1.utilities.time_utils import now_ist


def _split_by_date(start: datetime, end: datetime):
    cur = start
    while cur < end:
        next_midnight = datetime.combine(cur.date() + timedelta(days=1), datetime.min.time())
        chunk_end = min(end, next_midnight)
        yield cur.date(), (chunk_end - cur).total_seconds() / 60
        cur = chunk_end


def compute_daily_utilisation(wb, plan_start, plan_end) -> list[tuple]:
    """Return [(date, util_pct), ...] for each day in the plan window.

    Per-machine util = busy_min / 1440 (full calendar day), capped at 100%.
    Date-level util = mean across the full fleet (all machines in the Machine
    Utilization sheet, idle ones included as 0%).

    Shared by capacity_writer (per-date rows) and kpi_writer (overall average)
    so both use the identical denominator and capping.
    """
    # Fleet — Machine Utilization stores IDs as int, Shift Schedule as str.
    ws_util = wb["Machine Utilization"]
    all_machines: set = set()
    for r in range(4, ws_util.max_row + 1):
        m = ws_util.cell(row=r, column=1).value
        if m is not None:
            all_machines.add(str(m))
    if not all_machines:
        raise PipelineError(
            "Machine Utilization sheet has no machines listed",
            stage="upload", status_code=412,
        )

    # Per-(date, machine) busy minutes. PRODUCTIVE time only — changeover and
    # mould-clean rows are skipped so utilization reflects "press actually
    # producing tyres" not "press occupied for any reason".
    ws = wb["Shift Schedule"]
    busy: dict = defaultdict(float)
    for r in range(4, ws.max_row + 1):
        machine = ws.cell(row=r, column=3).value
        sku     = ws.cell(row=r, column=4).value
        s       = ws.cell(row=r, column=5).value
        e       = ws.cell(row=r, column=6).value
        remarks = ws.cell(row=r, column=10).value or ""
        if machine is None or s is None or e is None or e <= s:
            continue
        # Skip changeover rows (SKUCode == 'CHANGEOVER') and mould-clean rows
        # (Remarks contains 'Clean' or 'CLEAN'). Case-insensitive match.
        sku_u = str(sku).upper() if sku else ""
        remarks_u = str(remarks).upper()
        if sku_u == "CHANGEOVER" or "CLEAN" in remarks_u or "CHANGEOVER" in remarks_u:
            continue
        for d, mins in _split_by_date(s, e):
            busy[(d, str(machine))] += mins

    n_machines = len(all_machines)

    result = []
    d = plan_start
    while d <= plan_end:
        utils = [min(busy[(d, m)] / 1440.0, 1.0) for m in all_machines]
        util_pct = round(sum(utils) / n_machines * 100, 2)
        result.append((d, util_pct))
        d += timedelta(days=1)
    return result


def upload(schedule_path: Path, plan_id: str, created_by: str, db_cfg: dict) -> None:
    plan_row = plan_params.fetch(db_cfg, plan_id)
    plan_start = plan_row["planStartDate"]
    plan_end   = plan_row["planEndDate"]
    if isinstance(plan_start, datetime): plan_start = plan_start.date()
    if isinstance(plan_end,   datetime): plan_end   = plan_end.date()

    wb = openpyxl.load_workbook(schedule_path, data_only=True)
    try:
        daily = compute_daily_utilisation(wb, plan_start, plan_end)
    finally:
        wb.close()                                       # release file handle

    now = now_ist()
    rows = [(plan_id, d, util_pct, now, created_by) for d, util_pct in daily]
    print(f"[upload:capacity] {len(rows)} dates (plan window {plan_start}..{plan_end})")

    conn = connect(db_cfg)
    try:
        cur = conn.cursor()
        # NOTE: actual DB column is `creatdBy` (typo), not `createdBy`.
        cur.executemany(
            """INSERT INTO jkt_plan_capacityUtilisation
                   (plan_id, date, capacityUtilisation, createdAt, creatdBy)
               VALUES (%s, %s, %s, %s, %s)""",
            rows,
        )
        conn.commit()
        print(f"[upload:capacity] inserted {cur.rowcount} rows into jkt_plan_capacityUtilisation")
    finally:
        cur.close()
        conn.close()
