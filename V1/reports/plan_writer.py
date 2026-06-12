"""Bulk-insert every row of the schedule's Shift Schedule sheet into jkt_plan."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

from V1.utilities.db import connect, safe_table
from V1.utilities.time_utils import now_ist

_BATCH = 1000

# Index of `qty` and `sku_code` in the tuple appended to `rows`.
_QTY_IDX = 7
_SKU_IDX = 1


def _round_sku_totals_up_to_even(rows: list[tuple]) -> int:
    """Plant produces tyres in even counts only. Walk the rows, group by SKU,
    and for any SKU whose total qty is odd, add +1 to the qty of that SKU's
    first slot. Returns the number of SKUs that got bumped.

    CHANGEOVER rows (qty=0) are excluded from the totals — they don't produce
    tyres. Demand is NOT changed; only the planned qty gets nudged up by 1.
    """
    sku_total: dict = {}
    sku_first: dict = {}
    for i, r in enumerate(rows):
        sku = r[_SKU_IDX]
        qty = r[_QTY_IDX] or 0
        if not sku:
            continue                              # truly empty rows
        if str(sku).upper() == "CHANGEOVER":
            continue                              # never a producible SKU
        # Real SKUs WITH qty=0 are still tracked so totals dict matches DB SUM(qty).
        sku_total[sku] = sku_total.get(sku, 0) + int(qty)
        sku_first.setdefault(sku, i)

    # Use the SAME helper kpi_writer uses, so both writers agree on "evenness".
    from V1.reports.kpi_writer import _round_up_to_even
    bumped = 0
    for sku, total in sku_total.items():
        even_total = _round_up_to_even(total)
        if even_total != total:                 # was odd → bump first slot by +1
            i = sku_first[sku]
            row = list(rows[i])
            row[_QTY_IDX] = (row[_QTY_IDX] or 0) + 1
            rows[i] = tuple(row)
            bumped += 1
    return bumped


def _load_sku_descriptions(
    plan_id: str, db_cfg: dict, demand_table: str = "jkt_demand"
) -> dict[str, str]:
    """Build a SKUCode → description lookup from the demand table.

    The v4 scheduler dropped SKU_Description from the Shift Schedule sheet,
    so we enrich from the demand table (which has it). CHANGEOVER rows and
    any SKU not in the demand table simply get None.
    """
    demand_table = safe_table(demand_table)
    conn = connect(db_cfg)
    try:
        cur = conn.cursor()
        cur.execute(
            f"SELECT skuCode, MAX(skuDescription) FROM {demand_table} "
            "WHERE plan_id = %s GROUP BY skuCode",
            (plan_id,),
        )
        return {sku: desc for sku, desc in cur.fetchall() if sku}
    finally:
        cur.close()
        conn.close()


def upload(
    schedule_path: Path, plan_id: str, created_by: str, db_cfg: dict,
    tables: dict | None = None,
) -> None:
    tables          = tables or {}
    demand_table    = tables.get("demand", "jkt_demand")
    plan_table      = safe_table(tables.get("plan", "jkt_plan"))
    sku_desc_lookup = _load_sku_descriptions(plan_id, db_cfg, demand_table)

    wb = openpyxl.load_workbook(schedule_path, data_only=True)
    try:
        ws = wb["Shift Schedule"]
        now = now_ist()

        rows = []
        # Shift Schedule columns (v4): Date, Shift, Machine, SKUCode, StartTime,
        # EndTime, Qty, CycleTime_min, GT_Inventory, Remarks. Description is gone.
        for r in range(4, ws.max_row + 1):
            date_v, shift_v, _machine, sku_code, start_t, end_t, qty, cycle, _gt, remarks = (
                ws.cell(row=r, column=c).value for c in range(1, 11)
            )
            if all(v is None for v in (date_v, shift_v, sku_code, start_t, end_t, qty)):
                continue
            rows.append((
                plan_id,
                sku_code,
                sku_desc_lookup.get(sku_code),   # populated from jkt_demand
                date_v.date() if hasattr(date_v, "date") else date_v,
                shift_v,
                start_t,
                end_t,
                int(qty) if qty is not None else None,
                float(cycle) if cycle is not None else None,
                remarks,
                now,
                created_by,
            ))
    finally:
        wb.close()                                       # release file handle

    # Plant constraint: per-SKU planned qty must be EVEN. Add +1 tyre on the
    # first slot of any SKU whose total is odd.
    bumped = _round_sku_totals_up_to_even(rows)
    if bumped:
        print(f"[upload:plan] bumped +1 tyre on {bumped} SKUs to make totals even")

    conn = connect(db_cfg)
    try:
        cur = conn.cursor()
        total = 0
        for i in range(0, len(rows), _BATCH):
            cur.executemany(
                f"""INSERT INTO {plan_table}
                       (plan_id, skuCode, skuDescription, date, shift,
                        startTime, endTime, qty, cycleTime, remarks, createdAt, createdBy)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                rows[i:i + _BATCH],
            )
            total += cur.rowcount
        conn.commit()
        print(f"[upload:plan] inserted {total} rows into {plan_table}")
    finally:
        cur.close()
        conn.close()
