"""
Driver to run the OTHER developer's pipeline (bldg_curing2/curing_LP.py)
against the SAME Excel demand file that test_from_excel.py uses.

Steps it does for you:
1. Run our demand_route → compute CPS per SKU → produce requirement_summary.xlsx
2. Convert that to bldg_curing2/inputs/demand.csv (their format: SKUCode, Quantity, Priority)
3. Invoke their run_from_csv() to produce the 5-sheet schedule Excel

Usage:
    python3 test_from_excel_other.py --plan-id <id> \\
        --book "input/<demand>.xlsx" \\
        --sheet "<sheet name>" \\
        [--efficiency 94]

Output goes to: bldg_curing2/output/PCR_Curing_LP_v4_<plan_id>.xlsx
"""
from __future__ import annotations

import argparse
import csv
import math
import sys
from datetime import datetime
from pathlib import Path

ROOT  = Path(__file__).resolve().parent
OTHER = ROOT / "bldg_curing2"
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(OTHER))

import openpyxl

from V1.setups import demand_db, plan_params
from V1.routes.demand_route import _compute_cps, _resolve_weights
from V1.utilities import config_loader


def read_demand_from_excel(path: Path, sheet: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise SystemExit(f"Sheet {sheet!r} not in {path}; available: {wb.sheetnames}")
        ws = wb[sheet]
        rows: list[dict] = []
        skipped = 0
        for r in range(2, ws.max_row + 1):
            sku = ws.cell(row=r, column=1).value
            if not sku:
                continue
            req_raw = ws.cell(row=r, column=3).value
            try:
                if isinstance(req_raw, str):
                    cleaned = req_raw.strip().replace(",", "")
                    req = float(cleaned) if cleaned else 0.0
                else:
                    req = float(req_raw) if req_raw not in (None, "") else 0.0
                if not math.isfinite(req) or req < 0:
                    skipped += 1; continue
            except (ValueError, TypeError):
                skipped += 1; continue
            rows.append({
                "SKUCode":         sku,
                "SKU Description": ws.cell(row=r, column=2).value,
                "Requirement":     req,
                "Order Type":      ws.cell(row=r, column=4).value,
                "Market":          ws.cell(row=r, column=5).value,
                "Delivery date":   ws.cell(row=r, column=6).value,
            })
        if skipped:
            print(f"[driver] skipped {skipped} rows with non-numeric Requirement")
        return rows
    finally:
        wb.close()


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--plan-id", required=True,
                   help="plan_id in jkt_plan_params (used for weights / ranks / dates)")
    p.add_argument("--book", default="input/Book4.xlsx",
                   help="path to demand Excel")
    p.add_argument("--sheet", default="Sheet1", help="sheet name")
    p.add_argument("--efficiency", type=float, default=None,
                   help="override press_efficiency percent for this run (e.g., 94)")
    args = p.parse_args()

    # In-memory efficiency override (same trick as test_from_excel.py).
    if args.efficiency is not None:
        from V1.setups import plan_params as _pp
        _orig_fetch = _pp.fetch
        def _patched_fetch(db_cfg, plan_id):
            row = _orig_fetch(db_cfg, plan_id)
            row["efficiency"] = args.efficiency
            return row
        _pp.fetch = _patched_fetch
        print(f"[driver] efficiency override: {args.efficiency}%")

    cfg = config_loader.load()
    cfg["plan"]["plan_id"] = args.plan_id
    cfg = config_loader.resolve_paths(cfg)

    book_path = ROOT / args.book if not Path(args.book).is_absolute() else Path(args.book)
    print(f"[driver] plan_id={args.plan_id}  demand={book_path.name}  sheet={args.sheet!r}")

    # Step 1 — compute CPS using our Phase A.
    rows = read_demand_from_excel(book_path, args.sheet)
    if not rows:
        raise SystemExit("No demand rows found.")
    print(f"[driver] loaded {len(rows)} demand rows")

    plan_row = plan_params.fetch(cfg["db"], args.plan_id)
    weights = _resolve_weights(plan_row, cfg["demand"]["default_weights"])
    smin = int(cfg["demand"]["market_score_scale"]["min"])
    smax = int(cfg["demand"]["market_score_scale"]["max"])
    default_scores = cfg["demand"].get("default_market_scores")
    _compute_cps(rows, plan_row, weights,
                 cfg["demand"]["market_aliases"], smin, smax,
                 default_scores=default_scores)
    print("[driver] CPS computed")

    # Step 2 — write their demand.csv (SKUCode, Quantity, Priority).
    other_inputs = OTHER / "inputs"
    other_inputs.mkdir(parents=True, exist_ok=True)
    demand_csv = other_inputs / "demand.csv"
    with open(demand_csv, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["SKUCode", "Quantity", "Priority"])
        for r in rows:
            w.writerow([r["SKUCode"], int(r["Requirement"]),
                        round(r["ConsolidatedPriorityScore"], 6)])
    print(f"[driver] wrote {demand_csv}")

    # Step 3 — call their pipeline.
    print(f"[driver] running bldg_curing2/curing_LP.py via run_from_csv() ...")
    import curing_LP  # the other folder's module

    out_dir = OTHER / "output"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"OTHER_Schedule_{args.plan_id}.xlsx"

    # Plan start = jkt_plan_params.planStartDate at SHIFT_START_HOUR.
    ps = plan_row["planStartDate"]
    if isinstance(ps, datetime): ps = ps.date()
    plan_start = datetime(ps.year, ps.month, ps.day,
                          curing_LP.Config.SHIFT_START_HOUR, 0, 0)

    # Override planning days from DB.
    pe = plan_row["planEndDate"]
    if isinstance(pe, datetime): pe = pe.date()
    curing_LP.Config.PLANNING_DAYS = (pe - ps).days + 1
    if args.efficiency is not None:
        curing_LP.Config.PRESS_EFFICIENCY = args.efficiency / 100.0
    elif plan_row.get("efficiency"):
        curing_LP.Config.PRESS_EFFICIENCY = float(plan_row["efficiency"]) / 100.0
    if plan_row.get("noOfChangeOver") and int(plan_row["noOfChangeOver"]) > 0:
        curing_LP.Config.MAX_CHANGEOVERS_PER_SHIFT = int(plan_row["noOfChangeOver"])

    curing_LP.run_from_csv(
        input_dir=str(other_inputs),
        plan_start=plan_start,
        output_path=str(out_path),
    )
    print(f"\n[driver] DONE — schedule at {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
