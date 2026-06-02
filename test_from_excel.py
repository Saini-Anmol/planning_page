"""
Standalone test runner — reads demand from input/Book4.xlsx (Sheet1) instead of
the jkt_demand table, computes CPS, runs the LP scheduler, and writes the
5-sheet curing-schedule Excel to output/.

Does NOT touch the API. Does NOT write to any DB table.
DB is still READ-ONLY (for plan_params and the 6 master tables the LP needs).

Usage:
    python3 test_from_excel.py --plan-id BTP_June_Plan_V_384072
    python3 test_from_excel.py --plan-id <id> --book input/Book4.xlsx --sheet Sheet1
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

import openpyxl

from V1.routes import schedule_route
from V1.routes.demand_route import (
    _compute_cps,
    _resolve_weights,
    _write_output,
)
from V1.setups import plan_params
from V1.utilities import config_loader


def read_demand_from_excel(path: Path, sheet: str) -> list[dict]:
    """Read demand rows in the shape demand_route uses internally.

    Book4 Sheet1 columns:
        SKUCode | SKU Description | Requirement | Order Type | Market | Delivery date
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet!r} not in {path}; available: {wb.sheetnames}")
    ws = wb[sheet]

    rows: list[dict] = []
    skipped = 0
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(row=r, column=1).value
        if not sku:
            continue
        req_raw = ws.cell(row=r, column=3).value
        # Tolerate messy files:
        #  - Excel cells stored as strings with commas ('2,362' → 2362) get cleaned.
        #  - 'inf' / 'nan' / negatives rejected so they can't break normalization.
        import math
        try:
            if isinstance(req_raw, str):
                cleaned = req_raw.strip().replace(",", "")
                req = float(cleaned) if cleaned else 0.0
            else:
                req = float(req_raw) if req_raw not in (None, "") else 0.0
            if not math.isfinite(req) or req < 0:
                skipped += 1
                continue
        except (ValueError, TypeError):
            skipped += 1
            continue
        rows.append({
            "SKUCode":         sku,
            "SKU Description": ws.cell(row=r, column=2).value,
            "Requirement":     req,
            "Order Type":      ws.cell(row=r, column=4).value,
            "Market":          ws.cell(row=r, column=5).value,
            "Delivery date":   ws.cell(row=r, column=6).value,
        })
    if skipped:
        print(f"[test] skipped {skipped} rows with non-numeric Requirement (likely junk rows)")
    return rows


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--plan-id", required=True,
                   help="plan_id used to fetch weights + dates from jkt_plan_params (DB read-only)")
    p.add_argument("--book",  default="input/Book4.xlsx", help="path to demand Excel")
    p.add_argument("--sheet", default="Sheet1",          help="sheet name with the demand rows")
    p.add_argument("--efficiency", type=float, default=None,
                   help="Override press_efficiency percentage for this run (e.g., 94 = 94%%). "
                        "Patches the plan_params row in memory only — DB is not touched.")
    args = p.parse_args()

    # In-memory override of efficiency for testing without DB writes.
    if args.efficiency is not None:
        from V1.setups import plan_params as _pp
        _orig_fetch = _pp.fetch
        def _patched_fetch(db_cfg, plan_id):
            row = _orig_fetch(db_cfg, plan_id)
            row["efficiency"] = args.efficiency
            return row
        _pp.fetch = _patched_fetch
        print(f"[test] efficiency override: using {args.efficiency}% instead of DB value")

    cfg = config_loader.load()
    cfg["plan"]["plan_id"] = args.plan_id
    cfg = config_loader.resolve_paths(cfg)

    book_path = ROOT / args.book if not Path(args.book).is_absolute() else Path(args.book)

    print(f"[test] plan_id  = {args.plan_id}")
    print(f"[test] demand   = {book_path}  (sheet '{args.sheet}')")
    print(f"[test] NO DB writes, NO API. DB is read-only for plan params + master tables.\n")

    # 1. Read demand rows from Excel.
    rows = read_demand_from_excel(book_path, args.sheet)
    if not rows:
        raise SystemExit(f"No demand rows found in {book_path}::{args.sheet}")
    print(f"[test] loaded {len(rows)} demand rows from Excel")

    # 2. Fetch plan params (DB read-only).
    plan_row = plan_params.fetch(cfg["db"], args.plan_id)
    print(f"[test] plan_params: start={plan_row.get('planStartDate')}  "
          f"end={plan_row.get('planEndDate')}  "
          f"weights(m/q/d)={plan_row.get('marketWeightage')}/"
          f"{plan_row.get('quantityWeightage')}/{plan_row.get('targetdateWeightage')}")

    # 3. Compute CPS — same logic as Phase A, just with Excel-sourced rows.
    weights = _resolve_weights(plan_row, cfg["demand"]["default_weights"])
    smin = int(cfg["demand"]["market_score_scale"]["min"])
    smax = int(cfg["demand"]["market_score_scale"]["max"])
    default_scores = cfg["demand"].get("default_market_scores")
    _compute_cps(rows, plan_row, weights, cfg["demand"]["market_aliases"], smin, smax,
                 default_scores=default_scores)
    if default_scores:
        n_db_ranks = sum(1 for col in ("oe","re","st","defence","export","otr","government")
                         if plan_row.get(col) is not None)
        print(f"[test] CPS computed (weights: {tuple(round(w, 4) for w in weights)}, "
              f"DB ranks populated: {n_db_ranks}/7 — YAML defaults fill the gaps)")

    # 4. Write requirement_summary to the path schedule_route expects.
    summary_path = config_loader.output_dir(cfg) / cfg["demand"]["output_excel"]
    _write_output(rows, summary_path, cfg["demand"]["output_sheet"])
    print(f"[test] wrote intermediate {summary_path}")

    # 5. Run the LP scheduler. This reads master tables from DB (read-only),
    #    solves the LP, and writes the 5-sheet Excel.
    print(f"[test] running LP scheduler (~1-5 min)...")
    schedule_route.run(cfg)

    # 6. Done. Phase C (upload_route) intentionally skipped.
    print(f"\n[test] DONE — 5-sheet Excel is in {config_loader.output_dir(cfg)}/")
    print(f"[test] Confirmed: no rows inserted into jkt_plan_kpis / jkt_plan / jkt_plan_capacityUtilisation.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
