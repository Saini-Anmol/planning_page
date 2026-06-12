"""
building_schedule.py — JK Tyre PCR GREEN-TYRE BUILDING scheduler.

Builds the green tyres (GTs) that feed the curing schedule, synced to curing's
day-by-day consumption (JIT lead). Replaces the old static GT_Inventory: every
GT curing consumes must be produced here first.

Model (per the agreed spec)
  • Two production routes for a GT:
      - COMBINED  : one machine makes the GT directly.       ct = machine ct
      - TWO-STAGE : stage-1 makes the carcass, stage-2 makes the GT from it.
                    Treated as ONE process; scheduled on the STAGE-2 machine
                    with ct = max(stage1_ct, stage2_ct). Stage-1 carcass supply
                    is assumed non-binding (full component availability).
  • Demand per SKU = what the curing schedule actually plans to cure.
  • Sync = day-by-day JIT lead: cumulative GTs built must stay >= cumulative GTs
    cured (with a configurable BUILD_LEAD_DAYS head start). Capacity shortfalls
    are flagged, not hidden.
  • Same horizon / 3x8 shift structure as curing. Building cycle times are net
    (Norms ~= 480/ct), so no efficiency derate. No building changeover modelled.

Inputs (in ./inputs, via dataloader2.py)
    building_cycle_times.csv   SAPMachineCode, MachineName, StageName, Norms, CycleTime_min
    building_allowable.csv     SKUCode, Machines (list literal)
Plus the curing schedule (run live here) for the consumption profile.

Output
    output/PCR_Building_Schedule.xlsx   (6 sheets, same look as the curing book)

Run
    python building_schedule.py
"""

import ast
import math
import os
from collections import defaultdict
from datetime import datetime, timedelta

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from curing_LP import Config, ExcelExporter, _get_shift_fn
from curing_heuristic import run_heuristic_from_csv

# How many days ahead building must stay of curing (0 = same-day JIT).
BUILD_LEAD_DAYS = 1

# Building machine efficiency / OEE factor applied to cycle times.
# 1.0 = theoretical (Norms = 480/ct, no downtime derate) — per spec decision.
# Lower it (e.g. 0.90 to match curing, 0.85 realistic OEE) to derate capacity:
# effective ct = ct / BUILD_EFFICIENCY.
BUILD_EFFICIENCY = 1.0

# Changeover time charged when a building machine switches SKU.
#   "size"      : size-aware (default) — SameSize_Min if the two SKUs share rim
#                 size (chars 9-10 of the SKU code), else DifferentSize_Min.
#   "different" : always DifferentSize_Min  |  "same": always SameSize_Min
#   "none"      : disable changeover
BUILD_CO_MODE = "size"


def _sku_size(sku) -> str:
    """Rim diameter (inches) = characters 9-10 of the SKU code."""
    s = str(sku)
    return s[8:10] if len(s) >= 10 else ""

EVENT_SKUS = ["CHANGEOVER", "MOULD_CLEAN"]


# ══════════════════════════════════════════════════════════════════════════════
# CSV loaders for the building masters
# ══════════════════════════════════════════════════════════════════════════════
def load_building_cycle_times(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df["SAPMachineCode"] = df["SAPMachineCode"].astype(int)
    df["StageName"]      = df["StageName"].astype(str).str.strip().str.lower()
    df["CycleTime_min"]  = df["CycleTime_min"].astype(float)
    return df


def load_building_allowable(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype={"SKUCode": str})
    df["SKUCode"]  = df["SKUCode"].str.strip()
    df["Machines"] = df["Machines"].apply(
        lambda x: [int(m) for m in ast.literal_eval(x)]
        if isinstance(x, str) and x.strip() else []
    )
    return df


def load_building_changeover(path: str) -> dict:
    """Return {machine_code: {"same": min, "different": min}} from the masters."""
    if not os.path.exists(path):
        return {}
    df = pd.read_csv(path)
    return {int(r["MachineCode"]): {"same": float(r["SameSize_Min"]),
                                    "different": float(r["DifferentSize_Min"])}
            for _, r in df.iterrows()}


# ══════════════════════════════════════════════════════════════════════════════
# BUILDING SCHEDULER
# ══════════════════════════════════════════════════════════════════════════════
class BuildingScheduler:
    def __init__(self, plan_start: datetime, changeover: dict = None):
        self.plan_start  = plan_start
        self.shift_min   = Config.HOURS_PER_SHIFT * 60               # 480
        self.day_min     = Config.SHIFTS_PER_DAY * self.shift_min    # 1440
        self.changeover  = changeover or {}        # machine -> {"same","different"}

    def _co(self, machine, prev_sku, sku):
        """Changeover minutes for switching this machine from prev_sku to sku."""
        if prev_sku is None or prev_sku == sku or BUILD_CO_MODE == "none":
            return 0.0
        co = self.changeover.get(int(machine))
        if not co:
            return 0.0
        if BUILD_CO_MODE == "same":
            return co["same"]
        if BUILD_CO_MODE == "different":
            return co["different"]
        # size-aware: same rim size -> cheaper same-size changeover
        return co["same"] if _sku_size(prev_sku) == _sku_size(sku) else co["different"]

    # ── producer machines per SKU (with effective cycle time) ─────────────────
    def _build_producers(self, bld_ct: pd.DataFrame, bld_allow: pd.DataFrame):
        ct_map    = dict(zip(bld_ct["SAPMachineCode"], bld_ct["CycleTime_min"]))
        stage_map = dict(zip(bld_ct["SAPMachineCode"], bld_ct["StageName"]))
        name_map  = dict(zip(bld_ct["SAPMachineCode"], bld_ct["MachineName"]))

        producers: dict[str, list] = {}
        route:     dict[str, str]  = {}
        dropped_no_ct: set = set()

        for _, r in bld_allow.iterrows():
            sku = r["SKUCode"]
            allowed = r["Machines"]
            combined = [m for m in allowed if stage_map.get(m) == "combined"]
            stage1   = [m for m in allowed if stage_map.get(m) == "stage1"]
            stage2   = [m for m in allowed if stage_map.get(m) == "stage2"]
            dropped_no_ct |= {m for m in allowed if m not in stage_map}

            eff = BUILD_EFFICIENCY
            min_s1 = min((ct_map[m] for m in stage1), default=None)
            prod = []
            for m in combined:
                prod.append({"machine": m, "ct": ct_map[m] / eff, "route": "combined",
                             "name": name_map.get(m, str(m)), "stage": "combined"})
            if stage2 and min_s1 is not None:
                for m in stage2:
                    prod.append({"machine": m,
                                 "ct": max(ct_map[m], min_s1) / eff,
                                 "route": "two-stage",
                                 "name": name_map.get(m, str(m)),
                                 "stage": "stage2"})
            producers[sku] = prod

            has_c, has_2 = bool(combined), bool(stage2 and min_s1 is not None)
            route[sku] = ("both" if has_c and has_2 else
                          "combined" if has_c else
                          "two-stage" if has_2 else "none")

        return producers, route, ct_map, stage_map, name_map, dropped_no_ct

    # ── curing consumption profile (per SKU per day, cumulative) ───────────────
    @staticmethod
    def _curing_consumption(df_cure_shift: pd.DataFrame):
        prod = df_cure_shift[~df_cure_shift["SKUCode"].isin(EVENT_SKUS)].copy()
        prod["Date"] = pd.to_datetime(prod["Date"]).dt.date
        days = sorted(prod["Date"].unique())
        daily = (prod.groupby(["SKUCode", "Date"])["Qty"].sum()
                     .astype(int).to_dict())                       # (sku,date)->qty
        total = prod.groupby("SKUCode")["Qty"].sum().astype(int).to_dict()
        cured_per_day = prod.groupby("Date")["Qty"].sum().astype(int).to_dict()
        return days, daily, total, cured_per_day

    # ── JIT allocation, day by day ─────────────────────────────────────────────
    def allocate(self, producers, days, daily, priority):
        skus = sorted(producers.keys(), key=lambda s: -priority.get(s, 0.0))
        # cumulative curing demand per sku at each day index
        cum_need = defaultdict(dict)
        running = defaultdict(int)
        for d in days:
            for s in skus:
                running[s] += daily.get((s, d), 0)
                cum_need[s][d] = running[s]

        built_cum  = defaultdict(int)
        allocations = []                       # one row per (day, machine, sku)
        shortfalls  = []                       # (day, sku, missing_units)
        mlast: dict = {}                       # machine -> last SKU (across days)

        n = len(days)
        for i, day in enumerate(days):
            cap = defaultdict(lambda: self.day_min)         # fresh per day
            # JIT target: cover curing through (day + lead)
            tgt_idx = min(i + BUILD_LEAD_DAYS, n - 1)
            tgt_day = days[tgt_idx]
            for sku in skus:
                prod = producers.get(sku, [])
                if not prod:
                    continue
                need = cum_need[sku].get(tgt_day, 0) - built_cum[sku]
                if need <= 0:
                    continue
                # prefer a machine already on this SKU (no changeover), then
                # fastest machine, then most spare capacity
                order = sorted(prod, key=lambda p: (
                    0 if mlast.get(p["machine"]) == sku else 1,
                    p["ct"], -cap[p["machine"]]))
                for p in order:
                    if need <= 0:
                        break
                    m, ct = p["machine"], p["ct"]
                    co = self._co(m, mlast.get(m), sku)
                    if cap[m] - co < ct:
                        continue
                    take = min(need, int((cap[m] - co) / ct))
                    if take <= 0:
                        continue
                    cap[m]        -= take * ct + co
                    built_cum[sku] += take
                    need          -= take
                    mlast[m]       = sku
                    allocations.append({
                        "day": day, "machine": m, "sku": sku,
                        "units": take, "mins": round(take * ct, 2),
                        "co": round(co, 1),
                        "ct": ct, "route": p["route"],
                        "name": p["name"], "stage": p["stage"],
                    })
                if need > 0:
                    shortfalls.append((day, sku, int(need)))

        return allocations, built_cum, shortfalls

    # ── lay allocations out into shift-wise rows (with changeover rows) ─────────
    def build_shift_rows(self, allocations) -> pd.DataFrame:
        by_machine: dict = defaultdict(list)
        for a in allocations:
            by_machine[a["machine"]].append(a)

        rows = []
        for m, allocs in by_machine.items():
            allocs.sort(key=lambda a: a["day"])     # stable: keeps priority order within a day
            last_sku, cur_day, cursor = None, None, None
            for a in allocs:
                if a["units"] <= 0:
                    continue
                if a["day"] != cur_day:
                    cur_day = a["day"]
                    cursor = datetime(cur_day.year, cur_day.month, cur_day.day,
                                      Config.SHIFT_START_HOUR)
                # changeover row when the machine switches SKU
                if last_sku is not None and last_sku != a["sku"]:
                    co = self._co(m, last_sku, a["sku"])
                    if co > 0:
                        co_end = cursor + timedelta(minutes=co)
                        shift, _ = _get_shift_fn(cursor)
                        rows.append({
                            "Date": cur_day, "Shift": shift, "Machine": m,
                            "SKUCode": "CHANGEOVER", "StartTime": cursor,
                            "EndTime": co_end, "Qty": 0, "CycleTime_min": 0.0,
                            "Route": "", "Remarks": f"C/O {last_sku} -> {a['sku']}",
                        })
                        cursor = co_end
                # production block
                ct, units = a["ct"], a["units"]
                block_end = cursor + timedelta(minutes=a["mins"])
                inner, produced = cursor, 0
                while inner < block_end:
                    shift, shift_end = _get_shift_fn(inner)
                    slice_end = min(shift_end, block_end)
                    dur = (slice_end - inner).total_seconds() / 60
                    if dur <= 0:
                        inner = slice_end
                        continue
                    qty = (units - produced) if slice_end == block_end else int(dur / ct)
                    rows.append({
                        "Date": cur_day, "Shift": shift, "Machine": m,
                        "SKUCode": a["sku"], "StartTime": inner, "EndTime": slice_end,
                        "Qty": qty, "CycleTime_min": round(ct, 2),
                        "Route": a["route"], "Remarks": f"Building ({a['route']})",
                    })
                    produced += qty
                    inner = slice_end
                cursor = block_end
                last_sku = a["sku"]
        return pd.DataFrame(rows)


def load_building_masters(input_dir: str):
    return (load_building_cycle_times(os.path.join(input_dir, "building_cycle_times.csv")),
            load_building_allowable(os.path.join(input_dir, "building_allowable.csv")))


def building_capacity_envelope(input_dir: str, demand: dict, priority: dict) -> dict:
    """
    Max GTs the building fleet can make of each SKU over the whole horizon,
    allocated greedily in priority order across shared producer machines, with
    one changeover charged each time a machine takes on a new SKU. Returns
    {sku: deliverable_units}.  (Capacity ceiling used to cap curing.)
    """
    bld_ct, bld_allow = load_building_masters(input_dir)
    co_map = load_building_changeover(os.path.join(input_dir, "building_changeover.csv"))
    sched = BuildingScheduler(Config.PLAN_DATE, co_map)
    producers, route, *_ = sched._build_producers(bld_ct, bld_allow)
    horizon_min = Config.PLANNING_DAYS * sched.day_min

    cap_min = defaultdict(lambda: float(horizon_min))
    mlast: dict = {}
    # Seed EVERY demand SKU at 0 so SKUs building cannot make stay capped at 0
    # (curing must not cure a green tyre that cannot be built).
    supply: dict = {s: 0 for s in demand}
    for sku in sorted(demand, key=lambda s: -priority.get(s, 0.0)):
        need = int(demand[sku])
        if need <= 0:
            continue
        for p in sorted(producers.get(sku, []), key=lambda p: p["ct"]):
            if need <= 0:
                break
            m, ct = p["machine"], p["ct"]
            co = sched._co(m, mlast.get(m), sku)
            if cap_min[m] - co < ct:
                continue
            take = min(need, int((cap_min[m] - co) / ct))
            if take <= 0:
                continue
            cap_min[m]   -= take * ct + co
            mlast[m]      = sku
            supply[sku]  += take
            need         -= take
    return supply


# ══════════════════════════════════════════════════════════════════════════════
# ORCHESTRATION
# ══════════════════════════════════════════════════════════════════════════════
def run_building_from_csv(
    input_dir:   str = "inputs",
    plan_start:  datetime = None,
    output_dir:  str = "output",
    output_name: str = "PCR_Building_Schedule.xlsx",
    curing_results: dict = None,
    write_excel: bool = True,
) -> dict:
    if plan_start is None:
        plan_start = Config.PLAN_DATE

    # 1. Curing schedule = the consumption we must supply
    if curing_results is None:
        print("\n[Building] Running curing schedule for the consumption profile...")
        # write_excel=False: we only need curing's results in memory here, and we
        # must not clobber/lock an already-open curing workbook.
        curing_results = run_heuristic_from_csv(
            input_dir, plan_start, output_dir, write_excel=False)
    df_cure_shift = curing_results["shift_schedule"]
    cure_summary  = curing_results["demand_fulfillment"]
    priority = dict(zip(cure_summary["SKUCode"], cure_summary["Priority"]))

    # 2. Building masters
    bld_ct    = load_building_cycle_times(os.path.join(input_dir, "building_cycle_times.csv"))
    bld_allow = load_building_allowable(os.path.join(input_dir, "building_allowable.csv"))
    co_map    = load_building_changeover(os.path.join(input_dir, "building_changeover.csv"))

    print("\n" + "=" * 64)
    print("  JK Tyre PCR — GREEN-TYRE BUILDING Scheduler")
    print(f"  Plan start   : {plan_start:%Y-%m-%d %H:%M}")
    print(f"  Build lead   : {BUILD_LEAD_DAYS} day(s) (JIT)")
    print(f"  Machines     : {len(bld_ct)} "
          f"({(bld_ct['StageName']=='combined').sum()} combined, "
          f"{(bld_ct['StageName']=='stage1').sum()} stage1, "
          f"{(bld_ct['StageName']=='stage2').sum()} stage2)")
    print(f"  Changeover   : mode={BUILD_CO_MODE} | "
          f"{'set for '+str(len(co_map))+' machines' if co_map else 'none'}")
    print("=" * 64)

    sched = BuildingScheduler(plan_start, co_map)
    producers, route, ct_map, stage_map, name_map, dropped = \
        sched._build_producers(bld_ct, bld_allow)
    if dropped:
        print(f"  [Build] {len(dropped)} allowable machines have no cycle time "
              f"-> excluded: {sorted(dropped)}")

    days, daily, cure_total, cured_per_day = sched._curing_consumption(df_cure_shift)
    print(f"  [Build] Curing needs {sum(cure_total.values()):,} GTs across "
          f"{len(cure_total)} SKUs over {len(days)} days")

    allocations, built_cum, shortfalls = sched.allocate(producers, days, daily, priority)
    df_shift = sched.build_shift_rows(allocations)

    # 3. Reports
    results = _build_reports(
        sched, allocations, built_cum, shortfalls, df_shift,
        cure_total, daily, cured_per_day, days, producers, route,
        ct_map, stage_map, name_map, priority,
    )

    # 4. Export
    if write_excel:
        os.makedirs(output_dir, exist_ok=True)
        out_path = os.path.join(output_dir, output_name)
        BuildingExcelExporter(out_path).export(results)
    _print_building_summary(results, shortfalls)
    return results


def _build_reports(sched, allocations, built_cum, shortfalls, df_shift,
                   cure_total, daily, cured_per_day, days, producers, route,
                   ct_map, stage_map, name_map, priority):
    avail_total = len(days) * sched.day_min

    # ---- Machine Schedule ----
    agg: dict = defaultdict(lambda: {"units": 0, "mins": 0.0})
    meta: dict = {}
    for a in allocations:
        k = (a["machine"], a["sku"])
        agg[k]["units"] += a["units"]
        agg[k]["mins"]  += a["mins"]
        meta[k] = (a["name"], a["stage"], a["ct"], a["route"])
    mach_rows = []
    for (m, sku), v in agg.items():
        name, stage, ct, rt = meta[(m, sku)]
        mach_rows.append({
            "Machine": m, "MachineName": name, "Stage": stage,
            "SKUCode": sku, "Route": rt, "CycleTime_min": round(ct, 2),
            "Units_Built": v["units"], "Mins_Used": round(v["mins"], 1),
            "Days_Used": round(v["mins"] / sched.day_min, 2),
        })
    df_mach = pd.DataFrame(mach_rows).sort_values(["Machine", "SKUCode"]) \
        if mach_rows else pd.DataFrame(
            columns=["Machine", "MachineName", "Stage", "SKUCode", "Route",
                     "CycleTime_min", "Units_Built", "Mins_Used", "Days_Used"])

    # ---- GT Supply (fulfillment vs curing) ----
    sup_rows = []
    for sku, req in cure_total.items():
        built = int(built_cum.get(sku, 0))
        rt = route.get(sku, "none")
        nprod = len(producers.get(sku, []))
        if rt == "none":
            status, skip = "UNSCHEDULABLE", (
                "Not in building allowable" if sku not in route
                else "No combined and no stage1+stage2 route")
        elif built >= req:
            status, skip = "FULLY MET", ""
        elif built > 0:
            status, skip = "PARTIAL", "Building capacity short"
        else:
            status, skip = "UNMET", "Building capacity short"
        sup_rows.append({
            "SKUCode": sku, "Priority": round(priority.get(sku, 0.0), 4),
            "GT_Required": int(req), "GT_Built": built,
            "Gap": max(int(req) - built, 0),
            "Fulfillment_Pct": round(built / req * 100, 1) if req else 100.0,
            "Status": status, "Route": rt,
            "Producer_Machines": nprod, "Skip_Reason": skip,
        })
    df_supply = pd.DataFrame(sup_rows).sort_values("Priority", ascending=False)

    # ---- Machine Utilization ----
    used = defaultdict(float); units_m = defaultdict(int); skus_m = defaultdict(set)
    for a in allocations:
        used[a["machine"]]   += a["mins"]
        units_m[a["machine"]] += a["units"]
        skus_m[a["machine"]].add(a["sku"])
    util_rows = []
    for m in sorted(set(list(used.keys()) + [mm for mm in ct_map
                                             if stage_map.get(mm) in ("combined", "stage2")])):
        u = used.get(m, 0.0)
        util_rows.append({
            "Machine": m, "MachineName": name_map.get(m, str(m)),
            "Stage": stage_map.get(m, "?"),
            "Available_Mins": avail_total, "Used_Mins": round(u, 0),
            "Idle_Mins": round(avail_total - u, 0),
            "Utilization_Pct": round(u / avail_total * 100, 2) if avail_total else 0.0,
            "SKUs_Count": len(skus_m.get(m, set())),
            "Total_Units": units_m.get(m, 0),
        })
    df_util = pd.DataFrame(util_rows).sort_values("Utilization_Pct", ascending=False)

    # ---- Daily Building ----
    df_daily = ExcelExporter._daily_table(df_shift) if not df_shift.empty else \
        pd.DataFrame(columns=["Date", "Shift_A", "Shift_B", "Shift_C",
                              "Total_Units", "Cumulative_Units"])

    # ---- Sync (JIT) check ----
    built_per_day = (df_shift.groupby(pd.to_datetime(df_shift["Date"]).dt.date)["Qty"]
                     .sum().astype(int).to_dict()) if not df_shift.empty else {}
    sync_rows = []
    cb = cc = 0
    for d in days:
        b = int(built_per_day.get(d, 0)); c = int(cured_per_day.get(d, 0))
        cb += b; cc += c
        sync_rows.append({
            "Date": d, "Built_Today": b, "Cured_Today": c,
            "Cum_Built": cb, "Cum_Cured": cc, "Lead_Units": cb - cc,
            "Status": "OK" if cb >= cc else "LAG",
        })
    df_sync = pd.DataFrame(sync_rows)

    co_count = sum(1 for a in allocations if a.get("co", 0) > 0)
    co_mins  = sum(a.get("co", 0) for a in allocations)

    return {
        "gt_supply":           df_supply,
        "machine_schedule":    df_mach,
        "shift_schedule":      df_shift,
        "machine_utilization": df_util,
        "daily_building":      df_daily,
        "sync_check":          df_sync,
        "_meta": {
            "required": int(sum(cure_total.values())),
            "built":    int(sum(built_cum.values())),
            "days":     len(days),
            "shortfall_skus": len({s for _, s, _ in shortfalls}),
            "co_count": int(co_count),
            "co_hours": round(co_mins / 60, 1),
        },
    }


def _print_building_summary(results, shortfalls):
    m = results["_meta"]
    sup = results["gt_supply"]
    print(f"\n{'='*64}")
    print(f"  GT required (by curing): {m['required']:>10,}")
    print(f"  GT built               : {m['built']:>10,}  "
          f"({m['built']/m['required']*100:.1f}%)" if m['required'] else "")
    print(f"  Fully met SKUs         : {(sup['Status']=='FULLY MET').sum():>10}")
    print(f"  Partial SKUs           : {(sup['Status']=='PARTIAL').sum():>10}")
    print(f"  Unmet SKUs             : {(sup['Status']=='UNMET').sum():>10}")
    print(f"  No-build-route SKUs    : {(sup['Status']=='UNSCHEDULABLE').sum():>10}")
    lag = (results['sync_check']['Status'] == 'LAG').sum()
    print(f"  Days building lags curing: {lag:>8}")
    print(f"  Changeovers            : {m.get('co_count',0):>10,}  "
          f"({m.get('co_hours',0):,} hrs)")
    print(f"{'='*64}")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT  (reuses curing's styling helpers; building-specific sheets)
# ══════════════════════════════════════════════════════════════════════════════
class BuildingExcelExporter(ExcelExporter):
    def _sheet(self, writer, df, name, title, sub, widths,
               pct_cols=(), bold_cols=(), status_col=None, name_col=1,
               total_cols=None):
        df.to_excel(writer, sheet_name=name, index=False)
        ws = writer.book[name]
        ncol = len(df.columns)
        self._title(ws, title, sub, ncol)
        self._hdr_row(ws, 3, ncol)
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ri in range(4, ws.max_row + 1):
            stripe = "grey" if ri % 2 == 0 else "white"
            for ci in range(1, ncol + 1):
                fc = stripe
                if status_col and ci == status_col:
                    fc = self.STATUS_FC.get(str(ws.cell(ri, ci).value), stripe)
                self._cell(ws, ri, ci, ws.cell(ri, ci).value, fc=fc,
                           bold=(ci in bold_cols),
                           aln="left" if ci == name_col else "center")
            for pc in pct_cols:
                cell = ws.cell(ri, pc)
                if isinstance(cell.value, (int, float)):
                    cell.value = cell.value / 100
                cell.number_format = "0.0%"
        if total_cols:
            tr = ws.max_row + 1
            for ci in range(1, ncol + 1):
                c = ws.cell(tr, ci); c.fill = self.F("navy"); c.font = self._hf()
                c.border = self._b()
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(tr, 1).value = "TOTAL"
            for ci in total_cols:
                col = df.columns[ci - 1]
                ws.cell(tr, ci).value = int(df[col].sum())
                ws.cell(tr, ci).number_format = "#,##0"
        return ws

    def export(self, results: dict):
        m   = results["_meta"]
        req, built = m["required"], m["built"]
        pct = round(built / req * 100, 1) if req else 0.0
        lag = int((results["sync_check"]["Status"] == "LAG").sum())
        kpi = (f"GT required: {req:,}  |  GT built: {built:,}  |  "
               f"Coverage: {pct}%  |  Lead: {BUILD_LEAD_DAYS}d  |  "
               f"Lag days: {lag}  |  Changeovers: {m.get('co_count',0):,} "
               f"({m.get('co_hours',0):,} hrs)")

        with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
            # 1. GT Supply
            sup = results["gt_supply"][
                ["SKUCode", "Priority", "GT_Required", "GT_Built", "Gap",
                 "Fulfillment_Pct", "Status", "Route", "Producer_Machines",
                 "Skip_Reason"]]
            self._sheet(writer, sup, "GT Supply",
                        "GREEN-TYRE SUPPLY vs CURING DEMAND", kpi,
                        [26, 10, 13, 12, 10, 13, 14, 12, 14, 24],
                        pct_cols=(6,), bold_cols=(4,), status_col=7,
                        total_cols=(3, 4, 5))

            # 2. Machine Schedule
            mach = results["machine_schedule"]
            self._sheet(writer, mach, "Machine Schedule",
                        "BUILDING MACHINE-WISE SCHEDULE", kpi,
                        [12, 14, 12, 26, 12, 14, 14, 14, 12],
                        bold_cols=(1, 7), name_col=4)

            # 3. Shift Schedule
            shift = results["shift_schedule"]
            cols = ["Date", "Shift", "Machine", "SKUCode", "StartTime",
                    "EndTime", "Qty", "CycleTime_min", "Route", "Remarks"]
            shift = shift[cols] if not shift.empty else pd.DataFrame(columns=cols)
            self._sheet(writer, shift, "Shift Schedule",
                        "BUILDING SHIFT-WISE SCHEDULE", kpi,
                        [12, 8, 12, 26, 18, 18, 10, 12, 12, 22], name_col=4)

            # 4. Machine Utilization
            util = results["machine_utilization"]
            self._sheet(writer, util, "Machine Utilization",
                        "BUILDING MACHINE UTILIZATION", kpi,
                        [12, 14, 12, 15, 14, 14, 14, 12, 14],
                        pct_cols=(7,), bold_cols=(1, 7), name_col=2)

            # 5. Daily Building
            daily = results["daily_building"]
            self._sheet(writer, daily, "Daily Building",
                        "DAILY BUILDING — GTs BUILT PER DAY", kpi,
                        [14, 12, 12, 12, 14, 16],
                        bold_cols=(5,), name_col=1, total_cols=(2, 3, 4, 5))

            # 6. Sync (JIT) check
            sync = results["sync_check"]
            self._sheet(writer, sync, "Sync Check",
                        "JIT SYNC — BUILDING vs CURING (cumulative)", kpi,
                        [14, 13, 13, 14, 14, 12, 10],
                        bold_cols=(6,), status_col=7, name_col=1)

        print(f"\n  [Export] Saved -> {self.path}")


if __name__ == "__main__":
    run_building_from_csv("inputs", plan_start=Config.PLAN_DATE)
