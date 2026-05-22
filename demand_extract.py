"""
demand_extract.py — compute ConsolidatedPriorityScore (CPS) per SKU.

Reads:
  - plan parameters from jkplanningV1.jkt_plan_params (one row per plan_id)
  - per-SKU demand from an Excel sheet (default: Book4.xlsx / Sheet1)

Writes:
  - requirement_summary_<plan_id>.xlsx with columns
    SKUCode | SKU Description | Order Type | Market | Requirement | ConsolidatedPriorityScore

Configuration lives in demand_extract.yaml (DB creds, input/output paths,
market-name aliases, fallback weights, market score scale).

Usage:
    python demand_extract.py --plan-id 1001
    python demand_extract.py --plan-id 1001 --config demand_extract.yaml
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path
from typing import Any

import mysql.connector
import openpyxl
import yaml


DEFAULT_CONFIG: dict[str, Any] = {
    "db": {
        "host": "35.208.174.2", "port": 3306,
        "user": "root", "password": "Dev112233", "database": "jkplanningV1",
    },
    "input":  {"excel_path": "Book4.xlsx", "sheet": "Sheet1"},
    "output": {"excel_path": "requirement_summary_{plan_id}.xlsx",
               "sheet_name": "requirement_summary"},
    "market_aliases": {
        "OE": "oe", "Replacement": "re", "RE": "re", "ST": "st",
        "Defence": "defence", "Export": "export", "OTR": "otr",
        "Government": "government",
    },
    "default_weights": {"market": 0.50, "quantity": 0.20, "date": 0.30},
    "market_score_scale": {"min": 1, "max": 7},
}


def load_config(path: Path | None) -> dict:
    cfg = {k: (dict(v) if isinstance(v, dict) else v) for k, v in DEFAULT_CONFIG.items()}
    if path and path.exists():
        with open(path) as f:
            override = yaml.safe_load(f) or {}
        for k, v in override.items():
            if isinstance(v, dict) and isinstance(cfg.get(k), dict):
                cfg[k] = {**cfg[k], **v}
            else:
                cfg[k] = v
    return cfg


def fetch_plan_params(db_cfg: dict, plan_id: str) -> dict:
    conn = mysql.connector.connect(**db_cfg)
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM jkt_plan_params WHERE plan_id = %s", (plan_id,))
        row = cur.fetchone()
        if not row:
            raise SystemExit(f"plan_id={plan_id!r} not found in jkt_plan_params")
        return row
    finally:
        cur.close()
        conn.close()


EXPECTED_COLS = ["SKUCode", "SKU Description", "Requirement",
                 "Order Type", "Market", "Delivery date"]


def read_demand(path: Path, sheet: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet!r} not in {path}; available: {wb.sheetnames}")
    ws = wb[sheet]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}
    missing = [c for c in EXPECTED_COLS if c not in col]
    if missing:
        raise SystemExit(f"Demand sheet missing columns: {missing}. Found: {list(col)}")

    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(row=r, column=col["SKUCode"]).value
        if not sku:
            continue
        rows.append({
            "SKUCode":         sku,
            "SKU Description": ws.cell(row=r, column=col["SKU Description"]).value,
            "Requirement":     ws.cell(row=r, column=col["Requirement"]).value or 0,
            "Order Type":      ws.cell(row=r, column=col["Order Type"]).value,
            "Market":          ws.cell(row=r, column=col["Market"]).value,
            "Delivery date":   ws.cell(row=r, column=col["Delivery date"]).value,
        })
    return rows


def resolve_weights(plan_row: dict, defaults: dict) -> tuple[float, float, float]:
    """DB wins when non-null; fall back to YAML defaults. Renormalized to sum=1."""
    def pick(db_val, fallback):
        return float(db_val) if db_val not in (None, 0) else float(fallback)

    w_m = pick(plan_row.get("marketWeightage"),     defaults["market"])
    w_q = pick(plan_row.get("quantityWeightage"),   defaults["quantity"])
    w_d = pick(plan_row.get("targetdateWeightage"), defaults["date"])
    total = w_m + w_q + w_d
    if total == 0:
        raise SystemExit("All three weights are zero — cannot compute CPS")
    return w_m / total, w_q / total, w_d / total


def market_score(market_value: Any, plan_row: dict, aliases: dict, smin: int, smax: int) -> int:
    """Map free-text market → DB rank column → market_score (higher = higher priority).
    Falls back to smin when the market can't be resolved."""
    if not market_value:
        return smin
    key = str(market_value).strip()
    col = (aliases.get(key) or aliases.get(key.title())
           or aliases.get(key.upper()) or aliases.get(key.lower()))
    if not col:
        return smin
    rank = plan_row.get(col)
    if rank is None:
        return smin
    # rank is 1..N where 1 = highest priority. Invert so higher score = higher priority.
    rank = max(smin, min(smax, int(rank)))
    return smax + smin - rank


def _as_date(v: Any) -> dt.date | None:
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def compute_cps(rows: list[dict], plan_row: dict, weights: tuple[float, float, float],
                aliases: dict, smin: int, smax: int) -> None:
    plan_start = _as_date(plan_row.get("planStartDate"))
    plan_end   = _as_date(plan_row.get("planEndDate"))
    horizon = (plan_end - plan_start).days if (plan_start and plan_end) else 0

    for r in rows:
        r["_market_score"] = market_score(r["Market"], plan_row, aliases, smin, smax)

    # Missing target date → treat as plan end (least urgent).
    for r in rows:
        td = _as_date(r["Delivery date"])
        r["_ttt"] = (td - plan_start).days if (td and plan_start) else horizon

    reqs = [float(r["Requirement"] or 0) for r in rows]
    req_min, req_max = (min(reqs), max(reqs)) if reqs else (0.0, 0.0)
    ttts = [r["_ttt"] for r in rows]
    ttt_min, ttt_max = (min(ttts), max(ttts)) if ttts else (0, 0)

    def norm(v, lo, hi):
        return 0.0 if hi == lo else (v - lo) / (hi - lo)

    w_m, w_q, w_d = weights
    for r in rows:
        n_m = norm(r["_market_score"], smin, smax)
        n_q = norm(float(r["Requirement"] or 0), req_min, req_max)
        n_d = 0.0 if ttt_max == ttt_min else (ttt_max - r["_ttt"]) / (ttt_max - ttt_min)
        r["ConsolidatedPriorityScore"] = w_m * n_m + w_q * n_q + w_d * n_d


OUTPUT_COLS = ["SKUCode", "SKU Description", "Order Type", "Market",
               "Requirement", "ConsolidatedPriorityScore"]


def write_output(rows: list[dict], path: Path, sheet_name: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(OUTPUT_COLS)
    for r in rows:
        ws.append([r.get(c) for c in OUTPUT_COLS])
    wb.save(path)


def main() -> int:
    here = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Extract demand + compute CPS.")
    parser.add_argument("--plan-id", required=True, help="Primary key in jkt_plan_params")
    parser.add_argument("--config",  default=str(here / "demand_extract.yaml"),
                        help="Path to YAML config (default: demand_extract.yaml next to this script)")
    parser.add_argument("--input",   help="Override input Excel path")
    parser.add_argument("--output",  help="Override output Excel path")
    args = parser.parse_args()

    cfg = load_config(Path(args.config))

    plan_row = fetch_plan_params(cfg["db"], args.plan_id)
    print(f"Fetched plan {args.plan_id}: "
          f"start={plan_row.get('planStartDate')}  end={plan_row.get('planEndDate')}  "
          f"weights(market/qty/date)={plan_row.get('marketWeightage')}/"
          f"{plan_row.get('quantityWeightage')}/{plan_row.get('targetdateWeightage')}")

    weights = resolve_weights(plan_row, cfg["default_weights"])
    print(f"Renormalized weights (market, qty, date): "
          f"({weights[0]:.4f}, {weights[1]:.4f}, {weights[2]:.4f})")

    in_path = Path(args.input or cfg["input"]["excel_path"])
    if not in_path.is_absolute():
        in_path = here / in_path
    rows = read_demand(in_path, cfg["input"]["sheet"])
    print(f"Loaded {len(rows)} demand rows from {in_path.name}")

    smin = int(cfg["market_score_scale"]["min"])
    smax = int(cfg["market_score_scale"]["max"])
    compute_cps(rows, plan_row, weights, cfg["market_aliases"], smin, smax)

    out_template = args.output or cfg["output"]["excel_path"]
    out_path = Path(out_template.format(plan_id=args.plan_id))
    if not out_path.is_absolute():
        out_path = here / out_path
    write_output(rows, out_path, cfg["output"]["sheet_name"])
    print(f"Wrote {out_path}")

    for r in rows[:5]:
        print(f"  {r['SKUCode']}  market={r['Market']!r:14s}  "
              f"req={r['Requirement']!s:>6s}  CPS={r['ConsolidatedPriorityScore']:.6f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
